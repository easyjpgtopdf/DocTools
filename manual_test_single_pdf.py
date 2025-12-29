"""
Manual test of a single PDF to debug blank Excel issue
"""
import requests
from pathlib import Path
import json

# Configuration
BACKEND_URL = "https://pdf-to-excel-backend-564572183797.us-central1.run.app"
USER_ID = "NLhUrh6ZurQInLRV875Ktxw9rDn2"
USER_TYPE = "premium"

# Select a small PDF for testing
TEST_PDF = Path(r"C:\Users\apnao\Downloads\DocTools\Test Documents files\Aadhar.pdf")

print(f"Testing: {TEST_PDF.name}")
print(f"Size: {TEST_PDF.stat().st_size / 1024:.2f} KB")
print(f"Backend: {BACKEND_URL}")

# Convert
with open(TEST_PDF, 'rb') as f:
    files = {'file': (TEST_PDF.name, f, 'application/pdf')}
    headers = {
        'X-User-ID': USER_ID,
        'X-User-Type': USER_TYPE
    }
    
    print("\nSending request...")
    response = requests.post(
        f"{BACKEND_URL}/api/pdf-to-excel-docai",
        files=files,
        headers=headers,
        timeout=60
    )

print(f"\nStatus: {response.status_code}")

if response.status_code == 200:
    data = response.json()
    print("\n✅ SUCCESS")
    print(json.dumps(data, indent=2, ensure_ascii=False)[:1000])
    
    # Check key fields
    print("\n📊 KEY FIELDS:")
    print(f"  pages_processed (camelCase): {data.get('pagesProcessed', 'N/A')}")
    print(f"  execution_mode: {data.get('execution_mode', 'N/A')}")
    print(f"  mode: {data.get('mode', 'N/A')}")
    print(f"  layout_source: {data.get('layout_source', 'N/A')}")
    print(f"  engine: {data.get('engine', 'N/A')}")
    print(f"  downloadUrl: {data.get('downloadUrl', data.get('download_url', 'N/A'))[:80]}")
    print(f"  creditsDeducted: {data.get('creditsDeducted', 'N/A')}")
    
    # Download Excel
    excel_url = data.get('downloadUrl') or data.get('download_url')
    if excel_url:
        print(f"\n📥 Downloading Excel...")
        excel_resp = requests.get(excel_url, timeout=30)
        if excel_resp.status_code == 200:
            output_path = Path(r"C:\Users\apnao\Downloads\DocTools\Test Documents files\test_results\manual_test_output.xlsx")
            output_path.parent.mkdir(exist_ok=True)
            with open(output_path, 'wb') as f:
                f.write(excel_resp.content)
            
            print(f"✅ Excel downloaded: {output_path}")
            print(f"   Size: {len(excel_resp.content) / 1024:.2f} KB")
            
            # Analyze Excel
            import openpyxl
            wb = openpyxl.load_workbook(output_path, data_only=True)
            print(f"\n📊 Excel Analysis:")
            print(f"   Sheets: {len(wb.sheetnames)}")
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"   Sheet '{sheet_name}': {sheet.max_row} rows × {sheet.max_column} cols")
                
                # Show first few cells
                if sheet.max_row > 0:
                    print(f"      First cell (A1): '{sheet['A1'].value}'")
            wb.close()
else:
    print(f"\n❌ FAILED: {response.text[:500]}")

