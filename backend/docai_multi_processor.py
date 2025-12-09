"""
Multi-processor Document AI service supporting all processor types.
Route: /api/docai/process/:type
"""

import os
from google.cloud import documentai
from google.api_core import exceptions as gcp_exceptions
from typing import Tuple, Dict, Optional, List
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import logging
import uuid
from datetime import datetime

from storage_gcs import upload_excel_to_gcs, upload_pdf_to_gcs_temp, delete_from_gcs_temp

logger = logging.getLogger(__name__)

# Environment variables
DOCAI_PROJECT_ID = os.environ.get('DOCAI_PROJECT_ID', 'easyjpgtopdf-de346')
DOCAI_LOCATION = os.environ.get('DOCAI_LOCATION', 'us')
GCS_BUCKET = os.environ.get('GCS_BUCKET', '')

# Processor ID mapping
PROCESSOR_MAP: Dict[str, str] = {
    "form-parser-docai": os.environ.get('DOCAI_FORM_PARSER_ID', '9d1bf7e36946b781'),
    "layout-parser-docai": os.environ.get('DOCAI_LAYOUT_PARSER_ID', 'c79eead38f3ecc38'),
    "bank-docai": os.environ.get('DOCAI_BANK_ID', '6c8a0e5d0a3dddc4'),
    "expense-docai": os.environ.get('DOCAI_EXPENSE_ID', '3ac96b4ba3725046'),
    "identity-docai": os.environ.get('DOCAI_IDENTITY_ID', 'bd5e8109cd2ff2b9'),
    "pay-slip-docai": os.environ.get('DOCAI_PAY_SLIP_ID', '9034bca37aa74cff'),
    "utility-docai": os.environ.get('DOCAI_UTILITY_ID', '1c17582a99ad32d8'),
    "w2-docai": os.environ.get('DOCAI_W2_ID', '3090005e272cc32'),
    "w9-docai": os.environ.get('DOCAI_W9_ID', '91a5f06965a4a7a5'),
    # Default PDF to Excel processor
    "pdf-to-excel-docai": os.environ.get('DOCAI_PROCESSOR_ID', '19a07dc1c08ce733'),
}

# Initialize Document AI client
document_ai_client = None

def get_document_ai_client():
    """Initialize and return Document AI client using Application Default Credentials."""
    global document_ai_client
    if document_ai_client is None:
        document_ai_client = documentai.DocumentProcessorServiceClient()
    return document_ai_client

def get_processor_id(processor_type: str) -> Optional[str]:
    """
    Get processor ID for a given processor type.
    
    Args:
        processor_type: Processor type name (e.g., 'form-parser-docai')
    
    Returns:
        Processor ID or None if not found
    """
    return PROCESSOR_MAP.get(processor_type)

def get_available_processors() -> List[str]:
    """
    Get list of available processor types.
    
    Returns:
        List of processor type names that have valid processor IDs
    """
    return [ptype for ptype, pid in PROCESSOR_MAP.items() if pid and pid.strip()]

async def process_pdf_with_processor(
    file_bytes: bytes,
    filename: str,
    processor_type: str
) -> Tuple[str, int, int, Dict]:
    """
    Process PDF with specified Document AI processor and convert to Excel.
    
    Args:
        file_bytes: PDF file content as bytes
        filename: Original PDF filename
        processor_type: Type of processor to use (e.g., 'form-parser-docai')
    
    Returns:
        Tuple of (download_url, pages_processed, tables_found, extracted_data)
    
    Raises:
        ValueError: If processor type is invalid or processing fails
    """
    gcs_temp_uri = None
    try:
        # Validate processor type
        processor_id = get_processor_id(processor_type)
        if not processor_id:
            raise ValueError(f"Invalid processor type: {processor_type}. Available: {', '.join(get_available_processors())}")
        
        # Validate environment variables
        if not GCS_BUCKET:
            raise ValueError("GCS_BUCKET environment variable not set")
        if not DOCAI_PROJECT_ID:
            raise ValueError("DOCAI_PROJECT_ID environment variable not set")
        if not DOCAI_LOCATION:
            raise ValueError("DOCAI_LOCATION environment variable not set")
        
        # Step 1: Upload PDF to GCS (Document AI requires GCS input)
        logger.info(f"Uploading PDF to GCS for {processor_type} processing...")
        gcs_temp_uri = upload_pdf_to_gcs_temp(file_bytes, filename)
        logger.info(f"PDF uploaded to GCS: {gcs_temp_uri}")
        
        # Step 2: Build Document AI resource name
        processor_name = f"projects/{DOCAI_PROJECT_ID}/locations/{DOCAI_LOCATION}/processors/{processor_id}"
        
        # Step 3: Call Document AI
        logger.info(f"Processing document with {processor_type}: {processor_name}")
        client = get_document_ai_client()
        
        # Configure the process request
        from google.cloud.documentai_v1.types import document_processor_service
        
        gcs_document = document_processor_service.GcsDocument(
            gcs_uri=gcs_temp_uri,
            mime_type="application/pdf"
        )
        
        input_config = document_processor_service.InputConfig(
            gcs_document=gcs_document,
            mime_type="application/pdf"
        )
        
        request = document_processor_service.ProcessRequest(
            name=processor_name,
            input_config=input_config
        )
        
        # Process the document
        result = client.process_document(request=request)
        document = result.document
        
        # Step 4: Get page count
        pages_processed = len(document.pages)
        logger.info(f"Document processed. Pages: {pages_processed}")
        
        # Step 5: Extract data based on processor type
        logger.info(f"Extracting data from {processor_type} response...")
        extracted_data = extract_data_from_document(document, processor_type)
        tables_found = len(extracted_data.get('tables', []))
        logger.info(f"Found {tables_found} tables, {len(extracted_data.get('entities', []))} entities")
        
        # Step 6: Create Excel workbook
        logger.info("Creating Excel workbook...")
        excel_bytes = create_excel_from_extracted_data(extracted_data, filename, processor_type)
        logger.info(f"Excel created: {len(excel_bytes)} bytes")
        
        # Step 7: Upload Excel to GCS and get signed URL
        logger.info("Uploading Excel to GCS...")
        base_filename = filename.replace('.pdf', '').replace('.PDF', '')
        download_url = upload_excel_to_gcs(excel_bytes, f"{base_filename}_{processor_type}")
        logger.info(f"Excel uploaded. Download URL generated")
        
        return download_url, pages_processed, tables_found, extracted_data
        
    except gcp_exceptions.NotFound as e:
        logger.error(f"Document AI Processor not found or misconfigured: {e}")
        raise ValueError(f"Document AI Processor not found or misconfigured. Check processor ID for {processor_type}. Error: {e}")
    except gcp_exceptions.GoogleAPIError as e:
        logger.error(f"Google Document AI API error: {e}")
        raise ValueError(f"Google Document AI API error: {e}")
    except Exception as e:
        logger.error(f"Error in Document AI processing: {e}")
        raise ValueError(f"Document AI processing failed: {e}")
    finally:
        if gcs_temp_uri:
            logger.info(f"Deleting temporary PDF from GCS: {gcs_temp_uri}")
            delete_from_gcs_temp(gcs_temp_uri)

def extract_data_from_document(document, processor_type: str) -> Dict:
    """
    Extract data from Document AI response based on processor type.
    
    Args:
        document: Document AI Document object
        processor_type: Type of processor used
    
    Returns:
        Dictionary with extracted data (tables, entities, text, etc.)
    """
    extracted = {
        'tables': [],
        'entities': [],
        'text': document.text if hasattr(document, 'text') else '',
        'pages': len(document.pages),
        'processor_type': processor_type
    }
    
    # Extract tables from all pages
    for page in document.pages:
        if hasattr(page, 'tables') and page.tables:
            for table in page.tables:
                table_data = parse_docai_table(table, document.text if hasattr(document, 'text') else '')
                if table_data:
                    extracted['tables'].append(table_data)
    
    # Extract entities if available
    if hasattr(document, 'entities') and document.entities:
        for entity in document.entities:
            entity_data = {
                'type': entity.type_ if hasattr(entity, 'type_') else 'unknown',
                'value': entity.mention_text if hasattr(entity, 'mention_text') else '',
                'confidence': entity.confidence if hasattr(entity, 'confidence') else 0.0
            }
            extracted['entities'].append(entity_data)
    
    return extracted

def parse_docai_table(table, full_text: str) -> List[List[str]]:
    """
    Parse a Document AI table into a 2D list.
    """
    rows = []
    
    # Extract header rows
    if hasattr(table, 'header_rows') and table.header_rows:
        for header_row in table.header_rows:
            row_data = []
            for cell in header_row.cells:
                cell_text = extract_text_from_layout(cell.layout, full_text) if hasattr(cell, 'layout') else ''
                row_data.append(cell_text)
            if row_data:
                rows.append(row_data)
    
    # Extract body rows
    if hasattr(table, 'body_rows') and table.body_rows:
        for body_row in table.body_rows:
            row_data = []
            for cell in body_row.cells:
                cell_text = extract_text_from_layout(cell.layout, full_text) if hasattr(cell, 'layout') else ''
                row_data.append(cell_text)
            if row_data:
                rows.append(row_data)
    
    return rows if rows else None

def extract_text_from_layout(layout, full_text: str) -> str:
    """
    Extract text from a layout element.
    """
    if not layout or not hasattr(layout, 'text_anchor'):
        return ''
    
    try:
        text_anchor = layout.text_anchor
        if hasattr(text_anchor, 'text_segments') and text_anchor.text_segments:
            # Extract text from segments
            segments = []
            for segment in text_anchor.text_segments:
                if hasattr(segment, 'start_index') and hasattr(segment, 'end_index'):
                    start = segment.start_index if segment.start_index else 0
                    end = segment.end_index if segment.end_index else len(full_text)
                    segments.append(full_text[start:end])
            return ' '.join(segments).strip()
    except Exception as e:
        logger.warning(f"Error extracting text from layout: {e}")
    
    return ''

def create_excel_from_extracted_data(
    extracted_data: Dict,
    pdf_filename: str,
    processor_type: str
) -> bytes:
    """
    Create Excel workbook from extracted data.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    base_name = pdf_filename.replace('.pdf', '').replace('.PDF', '')
    
    # Add tables sheet
    tables = extracted_data.get('tables', [])
    if tables:
        for idx, table in enumerate(tables):
            sheet_name = f"Table {idx + 1}" if len(tables) > 1 else f"{processor_type[:31]}"
            sheet = workbook.create_sheet(title=sheet_name)
            
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    else:
        # Add entities sheet if no tables
        entities = extracted_data.get('entities', [])
        if entities:
            sheet = workbook.create_sheet(title=f"{processor_type[:31]}")
            sheet.cell(1, 1, "Type")
            sheet.cell(1, 2, "Value")
            sheet.cell(1, 3, "Confidence")
            
            for idx, entity in enumerate(entities, start=2):
                sheet.cell(idx, 1, entity.get('type', ''))
                sheet.cell(idx, 2, entity.get('value', ''))
                sheet.cell(idx, 3, entity.get('confidence', 0))
        else:
            sheet = workbook.create_sheet(title="No Data")
            sheet['A1'] = "No tables or entities detected in this PDF."
            sheet['A1'].font = Font(italic=True, color="666666")
    
    # Auto-adjust column widths
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

