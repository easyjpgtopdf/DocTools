"""
Google Document AI service for extracting tables from PDF documents.
"""

import os
from google.cloud import documentai
from google.api_core import exceptions as gcp_exceptions
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# Initialize Document AI client
document_ai_client = None
PROCESSOR_ID = os.environ.get('DOCAI_PROCESSOR_ID', '')
PROJECT_ID = os.environ.get('DOCAI_PROJECT_ID', 'easyjpgtopdf-de346')
LOCATION = os.environ.get('DOCAI_LOCATION', 'us-central1')


def get_document_ai_client():
    """Initialize and return Document AI client."""
    global document_ai_client
    if document_ai_client is None:
        document_ai_client = documentai.DocumentProcessorServiceClient()
    return document_ai_client


def process_document_with_tables(file_content: bytes, mime_type: str = 'application/pdf') -> documentai.Document:
    """
    Process PDF document with Document AI to extract tables.
    
    Args:
        file_content: PDF file content as bytes
        mime_type: MIME type of the document (default: application/pdf)
    
    Returns:
        Document AI response containing document structure with tables
    
    Raises:
        Exception: If Document AI processing fails
    """
    try:
        if not PROCESSOR_ID:
            raise ValueError("DOCAI_PROCESSOR_ID environment variable not set")
        
        client = get_document_ai_client()
        
        # Construct the full processor name
        name = f"projects/{PROJECT_ID}/locations/{LOCATION}/processors/{PROCESSOR_ID}"
        
        # Configure the process request
        raw_document = documentai.RawDocument(
            content=file_content,
            mime_type=mime_type
        )
        
        request = documentai.ProcessRequest(
            name=name,
            raw_document=raw_document
        )
        
        # Process the document
        logger.info(f"Processing document with Document AI processor: {PROCESSOR_ID}")
        result = client.process_document(request=request)
        document = result.document
        
        logger.info(f"Document processed successfully. Pages: {len(document.pages)}")
        return document
        
    except gcp_exceptions.GoogleAPIError as e:
        raise Exception(f"Document AI processing failed: {str(e)}")
    except Exception as e:
        raise Exception(f"Failed to process document: {str(e)}")


def get_page_count(document: documentai.Document) -> int:
    """
    Get page count from Document AI response.
    
    Args:
        document: Document AI document object
    
    Returns:
        Number of pages in the document
    """
    return len(document.pages)


def parse_document_ai_tables(document: documentai.Document) -> List[List[List[str]]]:
    """
    Parse Document AI response to extract table data.
    Returns a list of tables, where each table is a list of rows,
    and each row is a list of cell values.
    
    Args:
        document: Document AI document object
    
    Returns:
        List of tables, each table is a 2D list (rows x columns)
    """
    tables = []
    
    # Iterate through all pages
    for page in document.pages:
        # Extract tables from this page
        if page.tables:
            for table in page.tables:
                parsed_table = parse_table_from_document_ai(table, document.text)
                if parsed_table:
                    tables.append(parsed_table)
    
    return tables


def parse_table_from_document_ai(table: documentai.Document.Page.Table, full_text: str) -> List[List[str]]:
    """
    Parse a single table from Document AI.
    
    Args:
        table: Table object from Document AI
        full_text: Full text content of the document
    
    Returns:
        2D list representing the table (rows x columns)
    """
    if not table.header_rows or not table.body_rows:
        return []
    
    # Get table dimensions
    num_rows = len(table.header_rows) + len(table.body_rows)
    num_cols = 0
    
    # Find max columns
    for row in table.header_rows + table.body_rows:
        if row.cells:
            num_cols = max(num_cols, len(row.cells))
    
    if num_cols == 0:
        return []
    
    # Initialize grid
    grid = [['' for _ in range(num_cols)] for _ in range(num_rows)]
    
    # Process header rows
    row_idx = 0
    for header_row in table.header_rows:
        if header_row.cells:
            for col_idx, cell in enumerate(header_row.cells):
                if col_idx < num_cols:
                    cell_text = extract_text_from_layout(cell.layout, full_text)
                    grid[row_idx][col_idx] = cell_text
        row_idx += 1
    
    # Process body rows
    for body_row in table.body_rows:
        if body_row.cells:
            for col_idx, cell in enumerate(body_row.cells):
                if col_idx < num_cols:
                    cell_text = extract_text_from_layout(cell.layout, full_text)
                    grid[row_idx][col_idx] = cell_text
        row_idx += 1
    
    return grid


def extract_text_from_layout(layout: documentai.Document.Page.Layout, full_text: str) -> str:
    """
    Extract text from a layout element.
    
    Args:
        layout: Layout object from Document AI
        full_text: Full text content of the document
    
    Returns:
        Extracted text
    """
    if not layout or not layout.text_anchor:
        return ''
    
    text_anchor = layout.text_anchor
    if not text_anchor.text_segments:
        return ''
    
    # Extract text segments
    text_parts = []
    for segment in text_anchor.text_segments:
        start_index = int(segment.start_index) if segment.start_index else 0
        end_index = int(segment.end_index) if segment.end_index else 0
        
        if start_index < len(full_text) and end_index <= len(full_text):
            text_parts.append(full_text[start_index:end_index])
    
    return ' '.join(text_parts).strip()


def create_excel_from_tables(tables: List[List[List[str]]], pdf_filename: str) -> bytes:
    """
    Create Excel workbook from parsed tables.
    Each table becomes a separate sheet.
    
    Args:
        tables: List of tables (each table is a 2D list)
        pdf_filename: Original PDF filename (for sheet naming)
    
    Returns:
        Excel file as bytes
    """
    workbook = Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    base_name = pdf_filename.replace('.pdf', '').replace('.PDF', '')
    
    # If no tables found, create a single sheet with a message
    if not tables:
        sheet = workbook.create_sheet(title="No Tables")
        sheet['A1'] = "No tables detected in this PDF."
        sheet['A1'].font = Font(italic=True, color="666666")
    else:
        # Create a sheet for each table
        for idx, table in enumerate(tables):
            sheet_name = f"Table {idx + 1}" if len(tables) > 1 else base_name[:31]  # Excel sheet name limit
            sheet = workbook.create_sheet(title=sheet_name)
            
            # Write table data to sheet
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Style header row (first row)
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    
    # Auto-adjust column widths
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def convert_pdf_to_excel(file_content: bytes, pdf_filename: str):
    """
    Complete workflow: Process PDF with Document AI and create Excel.
    
    Args:
        file_content: PDF file content as bytes
        pdf_filename: Original PDF filename
    
    Returns:
        Tuple of (excel_bytes, page_count, tables)
    """
    # Step 1: Process document with Document AI
    logger.info("Processing PDF with Google Document AI")
    document = process_document_with_tables(file_content)
    
    # Step 2: Get page count
    page_count = get_page_count(document)
    logger.info(f"PDF has {page_count} pages")
    
    # Step 3: Parse tables
    logger.info("Parsing tables from Document AI response")
    tables = parse_document_ai_tables(document)
    logger.info(f"Found {len(tables)} tables")
    
    # Step 4: Create Excel
    logger.info("Creating Excel workbook")
    excel_bytes = create_excel_from_tables(tables, pdf_filename)
    
    return excel_bytes, page_count, tables

