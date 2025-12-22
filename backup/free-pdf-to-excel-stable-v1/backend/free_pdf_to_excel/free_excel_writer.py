"""
Excel Writer for FREE version.
Generates editable XLSX with basic formatting.
"""

import logging
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    HAS_OPENPYXL = True
except ImportError as e:
    HAS_OPENPYXL = False
    logger.warning(f"openpyxl not available: {e}")


def create_excel_workbook(
    grid: List[List[Dict]],
    header_rows: List[int],
    rectangles: List[Dict],
    images: List[Dict],
    row_boundaries: List[float],
    column_boundaries: List[float],
    output_path: str
) -> bool:
    """
    Create Excel workbook from grid with formatting.
    
    Args:
        grid: 2D grid of cells
        header_rows: List of row indices that are headers
        rectangles: List of rectangle objects (for background colors)
        images: List of image objects (for logos)
        row_boundaries: Row Y coordinates
        column_boundaries: Column X coordinates
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl not available")
        return False
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Page1"
        
        # Write data
        for row_idx, row in enumerate(grid):
            for col_idx, cell in enumerate(row):
                if cell.get('merged', False) and not cell.get('text'):
                    continue  # Skip merged cells
                
                cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                ws[cell_address] = cell['text']
                
                # Apply formatting
                cell_obj = ws[cell_address]
                
                # Font
                font = Font(
                    name=cell.get('font_name', 'Arial'),
                    size=cell.get('font_size', 10),
                    bold=cell.get('is_bold', False) or (row_idx in header_rows)
                )
                cell_obj.font = font
                
                # Header background color (if rectangle aligns)
                if row_idx in header_rows:
                    # Check if rectangle aligns with this row
                    has_background = False
                    if rectangles and row_boundaries:
                        row_y = row_boundaries[row_idx] if row_idx < len(row_boundaries) else 0
                        for rect in rectangles:
                            if abs(rect['y0'] - row_y) < 20 or abs(rect['y1'] - row_y) < 20:
                                # Check if rectangle spans this column
                                if column_boundaries:
                                    col_x = column_boundaries[col_idx] if col_idx < len(column_boundaries) else 0
                                    if rect['x0'] <= col_x <= rect['x1']:
                                        has_background = True
                                        break
                    
                    if has_background:
                        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        cell_obj.fill = fill
                
                # Alignment
                cell_obj.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Basic borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell_obj.border = thin_border
        
        # Auto-size columns
        if grid and len(grid) > 0 and len(grid[0]) > 0:
            for col_idx in range(len(grid[0])):
                col_letter = get_column_letter(col_idx + 1)
                max_length = 0
                for row in grid:
                    if col_idx < len(row):
                        cell_text = str(row[col_idx].get('text', ''))
                        max_length = max(max_length, len(cell_text))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Place images (logos) as floating objects
        # Note: openpyxl image placement is limited, we'll add comments instead
        for img in images:
            # Find nearest cell
            if column_boundaries and row_boundaries:
                col_idx = 0
                min_col_dist = float('inf')
                for i, col_x in enumerate(column_boundaries):
                    dist = abs(img['x0'] - col_x)
                    if dist < min_col_dist:
                        min_col_dist = dist
                        col_idx = i
                
                row_idx = 0
                min_row_dist = float('inf')
                for i, row_y in enumerate(row_boundaries):
                    dist = abs(img['y0'] - row_y)
                    if dist < min_row_dist:
                        min_row_dist = dist
                        row_idx = i
                
                # Add comment indicating image presence
                if row_idx < len(grid) and col_idx < len(grid[row_idx]):
                    cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                    cell_obj = ws[cell_address]
                    if not cell_obj.comment:
                        from openpyxl.comments import Comment
                        cell_obj.comment = Comment("Image/Logo detected at this position", "PDF Converter")
                        # Light yellow background to indicate image
                        fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                        cell_obj.fill = fill
        
        wb.save(output_path)
        return True
    
    except Exception as e:
        logger.error(f"Error creating Excel workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

