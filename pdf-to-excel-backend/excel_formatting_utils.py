"""
Excel formatting utilities for overlay enhancement.
Provides functions for Unicode handling, image insertion, and table formatting.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from io import BytesIO
import base64
import logging

logger = logging.getLogger(__name__)


def ensure_unicode_text(cell_value):
    """
    Ensure cell value is properly handled as Unicode.
    Preserves all Unicode characters including Hindi, Arabic, Chinese, etc.
    
    Args:
        cell_value: The cell value (string, number, or None)
    
    Returns:
        Properly encoded Unicode string or original value
    """
    if cell_value is None:
        return None
    
    if isinstance(cell_value, (int, float)):
        return cell_value
    
    # Convert to string and ensure Unicode handling
    try:
        if isinstance(cell_value, bytes):
            return cell_value.decode('utf-8', errors='replace')
        return str(cell_value)
    except Exception as e:
        logger.warning(f"Unicode conversion warning for value: {e}")
        return str(cell_value)


def apply_unicode_font(cell, font_name='Calibri'):
    """
    Apply Unicode-friendly font to a cell.
    Calibri supports most Unicode scripts including Devanagari, Arabic, CJK.
    
    Args:
        cell: openpyxl cell object
        font_name: Font name (default: Calibri for Unicode support)
    """
    try:
        current_font = cell.font or Font()
        cell.font = Font(
            name=font_name,
            size=current_font.size if current_font.size else 11,
            bold=current_font.bold,
            italic=current_font.italic,
            color=current_font.color
        )
    except Exception as e:
        logger.warning(f"Font application warning: {e}")


def insert_image_from_docai(worksheet, image_data, anchor_cell='A1', max_width=200, max_height=200):
    """
    Insert an image from Document AI into Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        image_data: Image data as bytes or base64 string
        anchor_cell: Cell to anchor image (e.g., 'A1')
        max_width: Maximum image width in pixels
        max_height: Maximum image height in pixels
    
    Returns:
        True if successful, False otherwise
    """
    try:
        # Handle base64 encoded image data
        if isinstance(image_data, str):
            if image_data.startswith('data:image'):
                # Remove data URL prefix
                image_data = image_data.split(',', 1)[1]
            image_bytes = base64.b64decode(image_data)
        else:
            image_bytes = image_data
        
        # Create image object
        img = Image(BytesIO(image_bytes))
        
        # Resize if needed
        if img.width > max_width or img.height > max_height:
            ratio = min(max_width / img.width, max_height / img.height)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        
        # Add image to worksheet
        worksheet.add_image(img, anchor_cell)
        return True
    except Exception as e:
        logger.warning(f"Image insertion warning: {e}")
        return False


def enhance_table_formatting(worksheet, start_row=1, end_row=None, start_col=1, end_col=None):
    """
    Enhance table formatting for better readability.
    Applies borders, alignment, and improves header styling.
    
    Args:
        worksheet: openpyxl worksheet object
        start_row: Starting row (1-indexed)
        end_row: Ending row (None = auto-detect)
        start_col: Starting column (1-indexed)
        end_col: Ending column (None = auto-detect)
    """
    try:
        # Auto-detect dimensions if not provided
        if end_row is None:
            end_row = worksheet.max_row
        if end_col is None:
            end_col = worksheet.max_column
        
        # Define styles
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Apply formatting
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Apply border
                cell.border = border
                
                # Header row styling
                if row_idx == start_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Alternate row coloring for readability
                    if row_idx % 2 == 0:
                        cell.fill = alternate_fill
                    
                    # Text alignment
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Ensure Unicode font
                    apply_unicode_font(cell)
    except Exception as e:
        logger.warning(f"Table formatting enhancement warning: {e}")


def detect_document_type(worksheet):
    """
    Detect document type based on content patterns.
    Returns: 'bank_statement', 'invoice', 'table', or 'general'
    """
    try:
        # Check first few rows for patterns
        text_content = ""
        for row in worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True):
            row_text = ' '.join(str(cell) if cell else '' for cell in row).lower()
            text_content += row_text + ' '
        
        # Bank statement patterns
        bank_keywords = ['account', 'balance', 'transaction', 'debit', 'credit', 'date', 'description', 'amount']
        if any(keyword in text_content for keyword in bank_keywords):
            return 'bank_statement'
        
        # Invoice patterns
        invoice_keywords = ['invoice', 'bill to', 'ship to', 'total', 'subtotal', 'tax', 'item', 'quantity', 'price']
        if any(keyword in text_content for keyword in invoice_keywords):
            return 'invoice'
        
        # Table detection (has structured rows/columns)
        if worksheet.max_row > 3 and worksheet.max_column > 2:
            return 'table'
        
        return 'general'
    except Exception as e:
        logger.warning(f"Document type detection warning: {e}")
        return 'general'


def enhance_bank_statement(worksheet):
    """
    Apply bank statement specific enhancements.
    - Better date formatting
    - Number alignment for amounts
    - Improved column widths
    """
    try:
        # Find amount column (usually last or second last)
        max_col = worksheet.max_column
        amount_col = max_col  # Assume last column is amount
        
        # Apply number formatting to amount column
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            cell = row[amount_col - 1]  # Convert to 0-indexed
            if cell.value:
                try:
                    # Try to format as currency
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                except:
                    pass
        
        # Auto-adjust column widths
        for col_idx in range(1, max_col + 1):
            max_length = 0
            for row in worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row)):
                cell = row[col_idx - 1]
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[worksheet.cell(1, col_idx).column_letter].width = adjusted_width
    except Exception as e:
        logger.warning(f"Bank statement enhancement warning: {e}")


def enhance_invoice(worksheet):
    """
    Apply invoice specific enhancements.
    - Better item table formatting
    - Total row highlighting
    - Improved spacing
    """
    try:
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Find and highlight total row (usually last row with "total" keyword)
        for row_idx in range(max_row, max(1, max_row - 5), -1):
            row_text = ' '.join(str(worksheet.cell(row_idx, col).value or '') for col in range(1, max_col + 1)).lower()
            if 'total' in row_text:
                # Highlight total row
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.font = Font(bold=True)
                break
        
        # Enhance item table
        enhance_table_formatting(worksheet, start_row=1, end_row=max_row, start_col=1, end_col=max_col)
    except Exception as e:
        logger.warning(f"Invoice enhancement warning: {e}")

