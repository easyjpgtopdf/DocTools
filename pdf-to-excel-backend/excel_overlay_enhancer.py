"""
Excel Overlay Enhancer for Premium PDF to Excel Conversion.

This module provides optional enhancement layer that improves Excel output
without affecting core conversion logic, credit system, or processor selection.

Features:
- Unicode language preservation
- Image insertion from Document AI
- Enhanced table formatting
- Document-type specific improvements (bank statements, invoices)
- Better cell alignment and styling

This is an OPTIONAL enhancement - if it fails, the original Excel is used.
"""

import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import logging
from typing import Optional

# Import formatting utilities
from excel_formatting_utils import (
    ensure_unicode_text,
    apply_unicode_font,
    insert_image_from_docai,
    enhance_table_formatting,
    detect_document_type,
    enhance_bank_statement,
    enhance_invoice
)

logger = logging.getLogger(__name__)


def enhance_excel_with_overlay(
    excel_bytes: bytes,
    docai_document=None,
    tables: list = None
) -> bytes:
    """
    Enhance Excel workbook with overlay improvements.
    
    This function:
    1. Loads the Excel workbook from bytes
    2. Applies Unicode text handling
    3. Inserts images from Document AI if available
    4. Enhances table formatting
    5. Applies document-type specific improvements
    6. Returns enhanced Excel as bytes
    
    Args:
        excel_bytes: Original Excel file as bytes
        docai_document: Document AI document object (optional, for image extraction)
        tables: List of extracted tables (optional, for context)
    
    Returns:
        Enhanced Excel file as bytes, or original bytes if enhancement fails
    """
    try:
        # Load workbook from bytes
        workbook = load_workbook(BytesIO(excel_bytes))
        
        # Process each worksheet
        for sheet in workbook.worksheets:
            try:
                # Step 1: Ensure Unicode handling for all cells
                _apply_unicode_to_sheet(sheet)
                
                # Step 2: Insert images if Document AI document is available
                if docai_document:
                    _insert_docai_images(sheet, docai_document)
                
                # Step 3: Detect document type and apply specific enhancements
                doc_type = detect_document_type(sheet)
                
                if doc_type == 'bank_statement':
                    enhance_bank_statement(sheet)
                elif doc_type == 'invoice':
                    enhance_invoice(sheet)
                else:
                    # General table enhancement
                    enhance_table_formatting(sheet)
                
                # Step 4: Auto-adjust column widths with Unicode consideration
                _auto_adjust_columns(sheet)
                
            except Exception as sheet_error:
                logger.warning(f"Error enhancing sheet '{sheet.title}': {sheet_error}")
                # Continue with other sheets even if one fails
        
        # Save enhanced workbook to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.warning(f"Excel overlay enhancement failed, using original: {e}")
        # Return original bytes if enhancement fails
        return excel_bytes


def _apply_unicode_to_sheet(worksheet):
    """
    Apply Unicode handling to all cells in a worksheet.
    Ensures proper encoding and font support for all languages.
    """
    try:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Ensure Unicode text
                    cell.value = ensure_unicode_text(cell.value)
                    # Apply Unicode-friendly font
                    apply_unicode_font(cell)
    except Exception as e:
        logger.warning(f"Unicode application warning: {e}")


def _insert_docai_images(worksheet, docai_document):
    """
    Extract and insert images from Document AI document into worksheet.
    Looks for image blocks in Document AI response and inserts them.
    Document AI may provide images in various formats - we try multiple approaches.
    """
    try:
        # Check if document has pages
        if not hasattr(docai_document, 'pages'):
            return
        
        image_count = 0
        max_images_per_sheet = 5  # Limit to avoid overwhelming the sheet
        
        for page in docai_document.pages:
            if image_count >= max_images_per_sheet:
                break
                
            # Method 1: Check for images attribute
            if hasattr(page, 'images') and page.images:
                for image in page.images[:max_images_per_sheet]:
                    if image_count >= max_images_per_sheet:
                        break
                    try:
                        image_bytes = None
                        
                        # Try different image content access methods
                        if hasattr(image, 'content') and image.content:
                            image_bytes = image.content
                        elif hasattr(image, 'encoded_content') and image.encoded_content:
                            import base64
                            image_bytes = base64.b64decode(image.encoded_content)
                        elif hasattr(image, 'mime_type') and hasattr(image, 'content'):
                            # Some Document AI versions structure differently
                            image_bytes = image.content
                        
                        if image_bytes:
                            # Insert image (anchor to column after data)
                            anchor_col = chr(ord('A') + min(worksheet.max_column + 1, 25)) if worksheet.max_column < 25 else 'Z'
                            anchor_cell = f"{anchor_col}{image_count * 50 + 1}"
                            
                            if insert_image_from_docai(worksheet, image_bytes, anchor_cell):
                                image_count += 1
                    except Exception as img_error:
                        logger.warning(f"Image extraction warning: {img_error}")
                        continue
            
            # Method 2: Check for visual elements that might contain images
            # (Some Document AI processors return images differently)
            if hasattr(page, 'visual_elements'):
                for visual_elem in page.visual_elements[:max_images_per_sheet]:
                    if image_count >= max_images_per_sheet:
                        break
                    # Visual elements might contain image data - skip for now
                    # as structure varies by processor type
                    pass
                    
    except Exception as e:
        logger.warning(f"Document AI image insertion warning: {e}")


def _auto_adjust_columns(worksheet):
    """
    Auto-adjust column widths considering Unicode characters.
    Some Unicode characters take more space, so we account for that.
    """
    try:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    try:
                        # Calculate length considering Unicode
                        cell_str = str(cell.value)
                        # Approximate: some Unicode chars are wider
                        # Count wide characters (CJK, some symbols)
                        wide_chars = sum(1 for c in cell_str if ord(c) > 127)
                        length = len(cell_str) + (wide_chars * 0.5)  # Wide chars take ~1.5x space
                        max_length = max(max_length, length)
                    except:
                        pass
            
            # Set column width (min 10, max 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        logger.warning(f"Column auto-adjustment warning: {e}")

