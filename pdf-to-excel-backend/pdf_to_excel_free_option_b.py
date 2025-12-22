"""
FREE Server-Side PDF to Excel Converter (Option-B)
Visual + Data Hybrid Pipeline
NO Document AI, NO OCR - CPU-only processing
"""

import os
import logging
import io
import hashlib
import time
from typing import Dict, List, Tuple, Optional
from datetime import datetime, timedelta
import traceback

try:
    from PyPDF2 import PdfReader
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTTextContainer, LTChar, LTRect, LTLine, LTFigure, LTImage
    from pdfminer.geometry import bbox2str
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_PDF_LIBS = True
except ImportError as e:
    HAS_PDF_LIBS = False
    import logging
    logging.warning(f"PDF libraries not available: {e}")

logger = logging.getLogger(__name__)

# Abuse control storage (in-memory for now, can be moved to Redis/DB)
_abuse_control = {}

# Constants
MAX_FREE_FILE_SIZE = 2 * 1024 * 1024  # 2 MB
MAX_FREE_PAGES_PER_DAY = 10  # Updated: 10 pages per day
FREE_RESET_HOURS = 24


def generate_free_key(ip_address: str, user_agent: str = "", fingerprint: str = "") -> str:
    """Generate a unique key for abuse control."""
    combined = f"{ip_address}|{user_agent}|{fingerprint}"
    return hashlib.sha256(combined.encode()).hexdigest()[:16]


def check_abuse_limits(free_key: str) -> Tuple[bool, str]:
    """
    Check if user has exceeded free limits.
    Returns: (allowed, message)
    """
    now = time.time()
    
    if free_key not in _abuse_control:
        _abuse_control[free_key] = {
            'pages_used_today': 0,
            'last_reset_timestamp': now,
            'last_file_size': 0
        }
        return True, ""
    
    user_data = _abuse_control[free_key]
    
    # Reset if 24 hours passed
    if now - user_data['last_reset_timestamp'] >= FREE_RESET_HOURS * 3600:
        user_data['pages_used_today'] = 0
        user_data['last_reset_timestamp'] = now
    
    # Check page limit
    if user_data['pages_used_today'] >= MAX_FREE_PAGES_PER_DAY:
        return False, f"Daily limit reached. You can convert up to {MAX_FREE_PAGES_PER_DAY} pages per day. Upgrade to Pro for unlimited conversions."
    
    return True, ""


def record_usage(free_key: str, pages: int, file_size: int):
    """Record usage for abuse control."""
    if free_key not in _abuse_control:
        _abuse_control[free_key] = {
            'pages_used_today': 0,
            'last_reset_timestamp': time.time(),
            'last_file_size': 0
        }
    
    _abuse_control[free_key]['pages_used_today'] += pages
    _abuse_control[free_key]['last_file_size'] = file_size


def extract_text_with_coordinates_for_page(pdf_bytes: bytes, page_num: int) -> List[Dict]:
    """
    Extract text with coordinates from a specific PDF page.
    Returns list of text objects with x, y, width, height, text.
    """
    text_objects = []
    
    try:
        pdf_file = io.BytesIO(pdf_bytes)
        
        for page_idx, page_layout in enumerate(extract_pages(pdf_file)):
            if page_idx == page_num:
                for element in page_layout:
                    if isinstance(element, LTTextContainer):
                        # Extract text and position
                        text = element.get_text().strip()
                        if not text:
                            continue
                        
                        # Get bounding box
                        x0, y0, x1, y1 = element.bbox
                        width = x1 - x0
                        height = y1 - y0
                        
                        # Get font info
                        font_name = "Arial"
                        font_size = 10
                        is_bold = False
                        
                        for text_line in element:
                            if isinstance(text_line, LTChar):
                                font_name = text_line.fontname or "Arial"
                                font_size = text_line.size or 10
                                if 'bold' in font_name.lower() or 'Bold' in font_name:
                                    is_bold = True
                                break
                        
                        text_objects.append({
                            'text': text,
                            'x': x0,
                            'y': y0,
                            'width': width,
                            'height': height,
                            'font_name': font_name,
                            'font_size': font_size,
                            'is_bold': is_bold,
                            'page': page_idx
                        })
                break
    
    except Exception as e:
        logger.error(f"Error extracting text from page {page_num}: {e}")
        logger.error(traceback.format_exc())
    
    return text_objects


def extract_lines_and_rectangles_for_page(pdf_bytes: bytes, page_num: int) -> Tuple[List[Dict], List[Dict]]:
    """
    Extract lines and rectangles from a specific PDF page.
    Returns: (lines, rectangles)
    """
    lines = []
    rectangles = []
    
    try:
        pdf_file = io.BytesIO(pdf_bytes)
        
        for page_idx, page_layout in enumerate(extract_pages(pdf_file)):
            if page_idx == page_num:
                for element in page_layout:
                    if isinstance(element, LTLine):
                        # Extract line coordinates
                        x0, y0, x1, y1 = element.bbox
                        lines.append({
                            'x0': x0,
                            'y0': y0,
                            'x1': x1,
                            'y1': y1,
                            'page': page_idx,
                            'is_horizontal': abs(y1 - y0) < abs(x1 - x0),
                            'is_vertical': abs(x1 - x0) < abs(y1 - y0)
                        })
                    
                    elif isinstance(element, LTRect):
                        # Extract rectangle
                        x0, y0, x1, y1 = element.bbox
                        rectangles.append({
                            'x0': x0,
                            'y0': y0,
                            'x1': x1,
                            'y1': y1,
                            'width': x1 - x0,
                            'height': y1 - y0,
                            'page': page_idx
                        })
                break
    
    except Exception as e:
        logger.error(f"Error extracting lines/rectangles from page {page_num}: {e}")
    
    return lines, rectangles


def detect_table_grid(lines: List[Dict], rectangles: List[Dict], page_width: float, page_height: float) -> Tuple[List[float], List[float], float]:
    """
    Detect table grid from lines and rectangles.
    Returns: (column_boundaries, row_boundaries, confidence)
    """
    # Extract horizontal and vertical lines
    horizontal_lines = [l for l in lines if l.get('is_horizontal', False)]
    vertical_lines = [l for l in lines if l.get('is_vertical', False)]
    
    # Find column boundaries from vertical lines
    column_xs = set()
    for line in vertical_lines:
        x = (line['x0'] + line['x1']) / 2
        column_xs.add(x)
    
    # Find row boundaries from horizontal lines
    row_ys = set()
    for line in horizontal_lines:
        y = (line['y0'] + line['y1']) / 2
        row_ys.add(y)
    
    # If no lines, try rectangles
    if not column_xs and rectangles:
        for rect in rectangles:
            column_xs.add(rect['x0'])
            column_xs.add(rect['x1'])
            row_ys.add(rect['y0'])
            row_ys.add(rect['y1'])
    
    # Sort boundaries
    column_boundaries = sorted(column_xs) if column_xs else []
    row_boundaries = sorted(row_ys, reverse=True) if row_ys else []  # Top to bottom
    
    # Calculate confidence based on line count
    line_count = len(horizontal_lines) + len(vertical_lines)
    confidence = min(line_count / 10.0, 1.0) if line_count > 0 else 0.0
    
    return column_boundaries, row_boundaries, confidence


def snap_text_to_grid(text_objects: List[Dict], column_boundaries: List[float], row_boundaries: List[float]) -> List[List[Dict]]:
    """
    Snap text objects to grid cells.
    Returns: 2D array of cells (rows x columns)
    """
    if not column_boundaries or not row_boundaries:
        return []
    
    # Initialize grid
    grid = []
    for _ in row_boundaries:
        row = [{'text': '', 'font_name': 'Arial', 'font_size': 10, 'is_bold': False} for _ in column_boundaries]
        grid.append(row)
    
    # Snap each text object to nearest cell
    for obj in text_objects:
        # Find closest column
        col_idx = 0
        min_col_dist = float('inf')
        for i, col_x in enumerate(column_boundaries):
            dist = abs(obj['x'] - col_x)
            if dist < min_col_dist:
                min_col_dist = dist
                col_idx = i
        
        # Find closest row
        row_idx = 0
        min_row_dist = float('inf')
        for i, row_y in enumerate(row_boundaries):
            dist = abs(obj['y'] - row_y)
            if dist < min_row_dist:
                min_row_dist = dist
                row_idx = i
        
        # Assign to cell (merge if already has text)
        if row_idx < len(grid) and col_idx < len(grid[row_idx]):
            cell = grid[row_idx][col_idx]
            if cell['text']:
                cell['text'] += ' ' + obj['text']
            else:
                cell['text'] = obj['text']
                cell['font_name'] = obj.get('font_name', 'Arial')
                cell['font_size'] = obj.get('font_size', 10)
                cell['is_bold'] = obj.get('is_bold', False)
    
    return grid


def detect_header_rows(grid: List[List[Dict]], rectangles: List[Dict], row_boundaries: List[float]) -> List[int]:
    """
    Detect header rows (usually first row, or rows with background rectangles).
    Returns: List of row indices that are headers
    """
    header_rows = []
    
    if not grid:
        return header_rows
    
    # First row is often header
    header_rows.append(0)
    
    # Check if rectangles align with first row
    if rectangles and row_boundaries:
        first_row_y = row_boundaries[0] if row_boundaries else 0
        for rect in rectangles:
            # Check if rectangle aligns with first row
            if abs(rect['y0'] - first_row_y) < 20:  # Within 20px
                header_rows.append(0)
                break
    
    return header_rows


def create_excel_from_grid(grid: List[List[Dict]], header_rows: List[int], ws) -> bool:
    """
    Create Excel worksheet from grid with formatting.
    """
    try:
        # Write data
        for row_idx, row in enumerate(grid):
            for col_idx, cell in enumerate(row):
                cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                ws[cell_address] = cell['text']
                
                # Apply formatting
                cell_obj = ws[cell_address]
                
                # Font
                font = Font(
                    name=cell.get('font_name', 'Arial'),
                    size=cell.get('font_size', 10),
                    bold=cell.get('is_bold', False) or (row_idx in header_rows)
                )
                cell_obj.font = font
                
                # Header background color
                if row_idx in header_rows:
                    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    cell_obj.fill = fill
                
                # Alignment
                cell_obj.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Basic borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell_obj.border = thin_border
        
        # Auto-size columns
        for col_idx in range(len(grid[0]) if grid else 0):
            col_letter = get_column_letter(col_idx + 1)
            max_length = 0
            for row in grid:
                if col_idx < len(row):
                    cell_text = str(row[col_idx].get('text', ''))
                    max_length = max(max_length, len(cell_text))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating Excel worksheet: {e}")
        logger.error(traceback.format_exc())
        return False


def process_pdf_to_excel_free_option_b(
    pdf_bytes: bytes,
    filename: str,
    free_key: str,
    ip_address: str
) -> Tuple[Optional[str], int, float, str]:
    """
    Process PDF to Excel using FREE Option-B (visual + data hybrid).
    
    Returns: (excel_path, pages_processed, confidence, error_message)
    """
    if not HAS_PDF_LIBS:
        return None, 0, 0.0, "PDF processing libraries not available"
    
    try:
        # Check file size
        file_size = len(pdf_bytes)
        if file_size > MAX_FREE_FILE_SIZE:
            return None, 0, 0.0, f"File too large. Maximum {MAX_FREE_FILE_SIZE / 1024 / 1024}MB for free version."
        
        # Check abuse limits
        allowed, message = check_abuse_limits(free_key)
        if not allowed:
            return None, 0, 0.0, message
        
        # Get page count
        try:
            pdf_file = io.BytesIO(pdf_bytes)
            reader = PdfReader(pdf_file)
            total_pages = len(reader.pages)
            
            if total_pages == 0:
                return None, 0, 0.0, "PDF has no pages"
            
            # FREE version: process up to 10 pages (daily limit)
            # Check remaining pages in daily limit
            if free_key in _abuse_control:
                pages_used = _abuse_control[free_key].get('pages_used_today', 0)
                pages_remaining = MAX_FREE_PAGES_PER_DAY - pages_used
                if pages_remaining <= 0:
                    return None, 0, 0.0, f"Daily limit reached. You can convert up to {MAX_FREE_PAGES_PER_DAY} pages per day. Upgrade to Pro for unlimited conversions."
                pages_to_process = min(total_pages, pages_remaining)
            else:
                pages_to_process = min(total_pages, MAX_FREE_PAGES_PER_DAY)
            
        except Exception as e:
            logger.error(f"Error reading PDF: {e}")
            return None, 0, 0.0, "Invalid PDF file"
        
        # Create Excel workbook
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        wb = openpyxl.Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        total_confidence = 0.0
        pages_processed = 0
        
        # Process each page
        for page_idx in range(pages_to_process):
            try:
                # Get page dimensions
                page = reader.pages[page_idx]
                page_width = 612  # Default
                page_height = 792  # Default
                if hasattr(page, 'mediabox'):
                    page_width = float(page.mediabox.width)
                    page_height = float(page.mediabox.height)
                
                # STEP 1: Extract text with coordinates for this page
                page_text_objects = extract_text_with_coordinates_for_page(pdf_bytes, page_idx)
                
                if not page_text_objects:
                    continue  # Skip empty pages
                
                # STEP 2: Extract lines and rectangles for this page
                page_lines, page_rectangles = extract_lines_and_rectangles_for_page(pdf_bytes, page_idx)
                
                # STEP 3: Detect table grid for this page
                column_boundaries, row_boundaries, grid_confidence = detect_table_grid(
                    page_lines, page_rectangles, page_width, page_height
                )
                
                # If confidence is too low, try to infer from text positions
                if grid_confidence < 0.2 and not column_boundaries and not row_boundaries:
                    if page_text_objects:
                        x_positions = sorted(set([obj['x'] for obj in page_text_objects]))
                        column_boundaries = []
                        for x in x_positions:
                            if not column_boundaries or x - column_boundaries[-1] > 30:
                                column_boundaries.append(x)
                        y_positions = sorted(set([obj['y'] for obj in page_text_objects]), reverse=True)
                        row_boundaries = []
                        for y in y_positions:
                            if not row_boundaries or abs(y - row_boundaries[-1]) > 10:
                                row_boundaries.append(y)
                
                # STEP 4: Snap text to grid
                grid = snap_text_to_grid(page_text_objects, column_boundaries, row_boundaries)
                
                if not grid or not any(any(cell['text'] for cell in row) for row in grid):
                    continue  # Skip pages without tables
                
                # STEP 5: Detect header rows
                header_rows = detect_header_rows(grid, page_rectangles, row_boundaries)
                
                # STEP 6: Add to Excel workbook
                ws = wb.create_sheet(title=f"Page{page_idx + 1}")
                
                # Create worksheet from grid
                if create_excel_from_grid(grid, header_rows, ws):
                    total_confidence += grid_confidence
                    pages_processed += 1
                
            except Exception as e:
                logger.error(f"Error processing page {page_idx + 1}: {e}")
                logger.error(traceback.format_exc())
                continue  # Continue with next page
        
        if pages_processed == 0:
            return None, 0, 0.0, "Could not extract table structure from any page. Try Premium for better accuracy."
        
        # Save workbook
        try:
            wb.save(temp_path)
        except Exception as e:
            logger.error(f"Error saving Excel: {e}")
            return None, 0, 0.0, "Error creating Excel file"
        
        # Record usage
        record_usage(free_key, pages_processed, file_size)
        
        # Calculate final confidence (average)
        final_confidence = min((total_confidence / pages_processed) + 0.3, 1.0) if pages_processed > 0 else 0.5
        
        return temp_path, pages_processed, final_confidence, ""
    
    except Exception as e:
        logger.error(f"Error in free option B processing: {e}")
        logger.error(traceback.format_exc())
        return None, 0, 0.0, f"Processing error: {str(e)}"
