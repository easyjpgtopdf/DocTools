"""
Google Document AI service for extracting tables from PDF documents.
New endpoint: /api/pdf-to-excel-docai
"""

import os
# Lazy import for documentai to avoid startup errors
# from google.cloud import documentai
from google.api_core import exceptions as gcp_exceptions
from typing import Tuple, List
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import logging
import uuid
from datetime import datetime

from storage_gcs import upload_excel_to_gcs, upload_pdf_to_gcs_temp, delete_from_gcs_temp

logger = logging.getLogger(__name__)

# Environment variables
DOCAI_PROJECT_ID = os.environ.get('DOCAI_PROJECT_ID', 'easyjpgtopdf-de346')
DOCAI_LOCATION = os.environ.get('DOCAI_LOCATION', 'us')  # 'us', 'eu', or 'asia'
DOCAI_PROCESSOR_ID = os.environ.get('DOCAI_PROCESSOR_ID', '')  # e.g., '19a07dc1c08ce733'
GCS_BUCKET = os.environ.get('GCS_BUCKET', '')

# Initialize Document AI client
document_ai_client = None


def get_document_ai_client():
    """Initialize and return Document AI client using Application Default Credentials."""
    global document_ai_client
    if document_ai_client is None:
        # Lazy import to avoid startup errors
        try:
            from google.cloud import documentai
            document_ai_client = documentai.DocumentProcessorServiceClient()
        except ImportError as e:
            raise ImportError(f"Failed to import google.cloud.documentai: {e}. Please install google-cloud-documentai.")
    return document_ai_client


def upload_pdf_to_gcs_temp(file_content: bytes, filename: str) -> str:
    """
    Upload PDF to GCS temporarily for Document AI processing.
    Returns GCS URI.
    """
    from google.cloud import storage
    
    if not GCS_BUCKET:
        raise ValueError("GCS_BUCKET environment variable not set")
    
    # Generate unique GCS blob name
    unique_id = str(uuid.uuid4())
    timestamp = datetime.now().strftime('%Y%m%d')
    blob_name = f"temp-pdf-uploads/{timestamp}/{unique_id}.pdf"
    
    try:
        client = storage.Client(project=DOCAI_PROJECT_ID)
        bucket = client.bucket(GCS_BUCKET)
        blob = bucket.blob(blob_name)
        
        blob.upload_from_string(
            file_content,
            content_type='application/pdf'
        )
        
        # Return GCS URI
        return f"gs://{GCS_BUCKET}/{blob_name}"
    except Exception as e:
        raise Exception(f"Failed to upload PDF to GCS: {str(e)}")


async def process_pdf_to_excel_docai(
    file_bytes: bytes, 
    filename: str,
    user_wants_premium: bool = False
) -> Tuple[str, int, List]:
    """
    Process PDF with Google Document AI and convert to Excel.
    
    Args:
        file_bytes: PDF file content as bytes
        filename: Original PDF filename
        user_wants_premium: User explicitly enabled premium mode toggle (default: False)
    
    Returns:
        Tuple of (download_url, pages_processed, unified_layouts)
    
    Raises:
        Exception: If processing fails
    """
    try:
        # Validate environment variables
        if not DOCAI_PROCESSOR_ID:
            raise ValueError("DOCAI_PROCESSOR_ID environment variable not set")
        if not GCS_BUCKET:
            raise ValueError("GCS_BUCKET environment variable not set")
        
        # Step 1: Early document type detection for processor routing
        filename_lower = filename.lower()
        document_text_preview = ''
        
        # Try to extract text preview for keyword detection (first 10KB)
        # This is optional - if PyPDF2 is not available, we'll use filename-based detection only
        try:
            import PyPDF2
            from io import BytesIO
            pdf_reader = PyPDF2.PdfReader(BytesIO(file_bytes))
            if len(pdf_reader.pages) > 0:
                first_page_text = pdf_reader.pages[0].extract_text()[:10000]  # First 10KB
                document_text_preview = first_page_text.lower()
                logger.debug("Extracted text preview for document type detection")
        except ImportError:
            logger.debug("PyPDF2 not available, using filename-based detection only")
        except Exception as e:
            logger.debug(f"PDF text extraction failed (non-critical): {e}")
            # Continue with filename-based detection
        
        # Keywords for bills/receipts/invoices
        bill_keywords = ['bill', 'receipt', 'payment', 'transaction', 'invoice', 'fee', 'total amount', 
                        'amount', 'price', 'cost', 'charge', 'paid', 'balance', 'due', 'subtotal']
        
        is_bill_or_receipt = any(keyword in filename_lower or keyword in document_text_preview for keyword in bill_keywords)
        
        # Step 1.1: Select appropriate processor
        processor_id_to_use = DOCAI_PROCESSOR_ID
        processor_type = "pdf-to-excel-docai"
        
        if is_bill_or_receipt:
            # Use Form Parser for bills/receipts
            form_parser_id = os.environ.get('DOCAI_FORM_PARSER_ID', '')
            if form_parser_id:
                processor_id_to_use = form_parser_id
                processor_type = "form-parser-docai"
                logger.info(f"Detected bill/receipt document, using Form Parser processor: {processor_id_to_use}")
            else:
                logger.warning("Form Parser ID not configured, using default processor")
        else:
            logger.info(f"Using default processor: {processor_id_to_use}")
        
        # Step 2: Upload PDF to GCS (Document AI requires GCS input)
        logger.info("Uploading PDF to GCS for Document AI processing...")
        gcs_uri = upload_pdf_to_gcs_temp(file_bytes, filename)
        logger.info(f"PDF uploaded to GCS: {gcs_uri}")
        
        # Step 3: Build Document AI resource name
        processor_name = f"projects/{DOCAI_PROJECT_ID}/locations/{DOCAI_LOCATION}/processors/{processor_id_to_use}"
        
        # Step 4: Call Document AI
        logger.info(f"Processing document with Document AI processor: {processor_type} ({processor_name})")
        client = get_document_ai_client()
        
        # Configure the process request
        # Use the correct Document AI v1 API structure
        # ProcessRequest available fields: gcs_document, raw_document, inline_document
        # For GCS input, use gcs_document (NOT gcs_source)
        try:
            from google.cloud.documentai_v1.types.document_processor_service import ProcessRequest
            from google.cloud.documentai_v1.types import GcsDocument
            
            # Create GcsDocument
            gcs_document = GcsDocument(
                gcs_uri=gcs_uri,
                mime_type="application/pdf"
            )
            
            # Create ProcessRequest with name and gcs_document
            request = ProcessRequest(
                name=processor_name,
                gcs_document=gcs_document
            )
            
        except (TypeError, AttributeError, ImportError) as e:
            # Strategy 2: Try setting fields individually
            try:
                from google.cloud.documentai_v1.types.document_processor_service import ProcessRequest
                from google.cloud.documentai_v1.types import GcsDocument
                
                request = ProcessRequest()
                request.name = processor_name
                
                # Set gcs_document field directly
                gcs_document = GcsDocument(
                    gcs_uri=gcs_uri,
            mime_type="application/pdf"
        )
                if hasattr(request, 'gcs_document'):
                    request.gcs_document.CopyFrom(gcs_document)
                else:
                    # Try direct assignment
                    request.gcs_document = gcs_document
                
            except (AttributeError, ImportError) as e2:
                # Strategy 3: Try setting gcs_document fields directly
                try:
                    from google.cloud.documentai_v1.types.document_processor_service import ProcessRequest
                    
                    request = ProcessRequest()
                    request.name = processor_name
                    
                    # Try setting gcs_document fields directly
                    if hasattr(request, 'gcs_document'):
                        request.gcs_document.gcs_uri = gcs_uri
                        request.gcs_document.mime_type = "application/pdf"
                    else:
                        raise AttributeError(
                            f"ProcessRequest does not have 'gcs_document' field. "
                            f"Available fields: {dir(request)[:20]}"
                        )
                        
                except (AttributeError, ImportError) as e3:
                    raise ImportError(
                        f"Could not construct Document AI request. "
                        f"Error 1: {e}, Error 2: {e2}, Error 3: {e3}. "
                        f"Please check google-cloud-documentai installation and version."
        )
        
        # Process the document
        result = client.process_document(request=request)
        document = result.document
        
        # Step 5: Get page count
        pages_processed = len(document.pages)
        logger.info(f"Document processed. Pages: {pages_processed}")
        
        # Step 6: Extract tables and structure
        logger.info("Extracting tables from Document AI response...")
        logger.info(f"Processor used: {processor_type}")
        logger.info(f"Document type: {type(document)}")
        logger.info(f"Document has text: {hasattr(document, 'text') and bool(document.text)}")
        logger.info(f"Document pages: {len(document.pages) if hasattr(document, 'pages') else 0}")
        
        # Extract native table objects for premium layout system
        native_tables = []
        for page in document.pages:
            if hasattr(page, 'tables') and page.tables:
                native_tables.extend(page.tables)
        
        # Extract parsed tables for standard method
        tables = extract_tables_from_document(document)
        
        # Count form fields
        form_fields_count = 0
        for page in document.pages:
            if hasattr(page, 'form_fields') and page.form_fields:
                form_fields_count += len(page.form_fields)
        
        # Count blocks
        blocks_count = 0
        for page in document.pages:
            if hasattr(page, 'blocks') and page.blocks:
                blocks_count += len(page.blocks)
        
        logger.info(f"Tables found: {len(tables)} parsed, {len(native_tables)} native")
        logger.info(f"Form fields found: {form_fields_count}")
        logger.info(f"Blocks found: {blocks_count}")
        
        if not tables and not native_tables:
            logger.warning("WARNING: No tables extracted from Document AI response!")
            logger.warning("Will attempt structure reconstruction from form fields and blocks")
        
        # Step 7: Create Excel workbook
        logger.info("Creating Excel workbook...")
        # Pass document.text for fallback text extraction
        document_text = document.text if hasattr(document, 'text') else ''
        logger.info(f"Document text available: {len(document_text)} characters")
        
        # Step 7.0: Try premium layout system (MANDATORY - no text-only fallback)
        excel_bytes = None
        layout_strategy = "unknown"
        
        # Initialize unified_layouts to empty list (will be set in try block)
        unified_layouts = []
        excel_bytes = None
        layout_strategy = "unknown"
        decision_engine_instance = None  # CRITICAL: Initialize in function scope
        
        try:
            from premium_layout.layout_decision_engine import LayoutDecisionEngine
            from premium_layout.excel_word_renderer import ExcelWordRenderer
            
            logger.info("=" * 80)
            logger.info("PREMIUM PIPELINE: Starting document processing")
            logger.info("=" * 80)
            logger.info("Premium layout system: Starting enhanced extraction...")
            decision_engine = LayoutDecisionEngine()
            # CRITICAL: This call triggers:
            # 1. DecisionRouter selects ONE execution mode
            # 2. Execute ONLY the selected mode
            # 3. TablePostProcessor.process_table() - 7-step pipeline (for TABLE_STRICT)
            # 4. ProcessedTable -> UnifiedLayout conversion
            # STEP-13: Pass PDF bytes for Adobe fallback
            unified_layouts = decision_engine.process_document(
                document, 
                document_text, 
                native_tables, 
                processor_type=processor_type,
                pdf_bytes=file_bytes  # STEP-13: Pass PDF bytes for Adobe fallback
            )
            logger.info(f"LayoutDecisionEngine.process_document() returned {len(unified_layouts)} UnifiedLayout objects")
            
            # CRITICAL: Extract pages_metadata from decision_engine (STEP-14 builds it)
            # This is the source of truth for billing
            pages_metadata_from_engine = getattr(decision_engine, 'pages_metadata', None)
            if pages_metadata_from_engine:
                logger.critical(f"✅ STEP-14: Retrieved {len(pages_metadata_from_engine)} pages metadata from decision_engine")
                for pm in pages_metadata_from_engine:
                    logger.critical(f"   Page {pm.get('page')}: engine={pm.get('engine')}")
            else:
                logger.warning("⚠️ STEP-14: pages_metadata not found in decision_engine - will extract from unified_layouts")
            
            # Store decision engine for metadata extraction (ENTERPRISE RESPONSE)
            # This allows main.py to extract mode, confidence, message
            decision_engine_instance = decision_engine
            
            # STEP 6.5: ADOBE FALLBACK (SELECTIVE)
            # Check if Adobe PDF Extract should be used as fallback
            try:
                from premium_layout.decision_router import DecisionRouter
                from premium_layout.adobe_fallback_service import get_adobe_fallback_service
                
                # Extract DocAI confidence from unified_layouts metadata
                docai_confidence = 0.5
                if unified_layouts and len(unified_layouts) > 0:
                    first_layout = unified_layouts[0]
                    if hasattr(first_layout, 'metadata') and first_layout.metadata:
                        docai_confidence = first_layout.metadata.get('routing_confidence', 0.5)
                
                # Create full_structure dict for complexity assessment
                full_structure = {
                    'blocks': [],
                    'tables': native_tables or []
                }
                # Extract blocks from document pages
                for page in document.pages:
                    for block in page.blocks:
                        block_dict = {
                            'text': block.layout.text_anchor.content if hasattr(block.layout, 'text_anchor') else '',
                            'bounding_box': {}
                        }
                        if block.layout.bounding_poly:
                            vertices = block.layout.bounding_poly.normalized_vertices
                            if len(vertices) >= 4:
                                block_dict['bounding_box'] = {
                                    'x_min': min(v.x for v in vertices),
                                    'x_max': max(v.x for v in vertices),
                                    'y_min': min(v.y for v in vertices),
                                    'y_max': max(v.y for v in vertices)
                                }
                        full_structure['blocks'].append(block_dict)
                
                # =====================================================================
                # STEP 6.5: ADOBE PDF EXTRACT API FALLBACK (WITH STRICT GUARDRAILS)
                # =====================================================================
                
                # FEATURE FLAGS CHECK (FIRST)
                adobe_allowed_by_flags = True
                feature_flag_reason = ""
                try:
                    from feature_flags import get_feature_flags
                    flags = get_feature_flags()
                    
                    adobe_allowed_by_flags, feature_flag_reason = flags.can_use_adobe(
                        user_wants_premium=user_wants_premium,
                        docai_confidence=docai_confidence,
                        page_count=len(document.pages) if hasattr(document, 'pages') else 1
                    )
                    
                    if not adobe_allowed_by_flags:
                        logger.info(f"🚫 Adobe blocked by feature flags: {feature_flag_reason}")
                except ImportError:
                    logger.warning("Feature flags module not available - Adobe fallback will proceed")
                except Exception as ff_err:
                    logger.error(f"Feature flags check failed: {ff_err}")
                
                # Continue only if feature flags allow Adobe
                if adobe_allowed_by_flags:
                    router = DecisionRouter()
                    
                    # Extract page count for cost calculation
                    page_count = len(document.pages) if hasattr(document, 'pages') else 1
                    
                    # Apply comprehensive guardrails
                    should_fallback, fallback_reason, guardrail_metadata = router.should_enable_adobe_with_guardrails(
                        user_wants_premium=user_wants_premium,
                        docai_confidence=docai_confidence,
                        full_structure=full_structure,
                        unified_layouts=unified_layouts,
                        page_count=page_count,
                        user_plan="premium"
                    )
                    
                    # Log guardrail results (MANDATORY)
                    logger.info("=" * 80)
                    logger.info("ADOBE FALLBACK GUARDRAILS CHECK")
                    logger.info(f"Adobe fallback allowed: {'YES' if should_fallback else 'NO'}")
                    logger.info(f"Reason: {fallback_reason}")
                    logger.info(f"Gates passed: {', '.join(guardrail_metadata.get('gates_passed', []))}")
                    logger.info(f"Gates failed: {', '.join(guardrail_metadata.get('gates_failed', []))}")
                    logger.info(f"Estimated Adobe cost: {guardrail_metadata.get('estimated_cost_credits', 0)} credits for {page_count} pages")
                    logger.info("=" * 80)
                else:
                    should_fallback = False
                    fallback_reason = f"Feature flags blocked Adobe: {feature_flag_reason}"
                    guardrail_metadata = {
                        'gates_passed': [],
                        'gates_failed': ['feature_flags_blocked'],
                        'estimated_adobe_pages': 0,
                        'estimated_cost_credits': 0
                    }
                
                if should_fallback:
                    adobe_service = get_adobe_fallback_service()
                    
                    if adobe_service.is_available():
                        logger.info("=" * 80)
                        logger.info("🚀 ADOBE FALLBACK: Starting extraction")
                        logger.info(f"Structural failures: {', '.join(guardrail_metadata.get('failure_reasons', []))}")
                        logger.info(f"Document: {filename} ({page_count} pages)")
                        logger.info("=" * 80)
                        
                        # Call Adobe PDF Extract API
                        adobe_layouts, adobe_confidence, adobe_metadata = adobe_service.extract_pdf_structure(
                            pdf_bytes=file_bytes,
                            filename=filename,
                            docai_confidence=docai_confidence
                        )
                        
                        # Replace layouts ONLY if Adobe succeeded
                        if adobe_layouts and len(adobe_layouts) > 0:
                            logger.info(f"✅ Adobe SUCCESS: Replacing {len(unified_layouts)} DocAI layouts with {len(adobe_layouts)} Adobe layouts")
                            unified_layouts = adobe_layouts
                            
                            # Update metadata to indicate Adobe was used (CRITICAL for billing)
                            for idx, layout in enumerate(unified_layouts):
                                if hasattr(layout, 'metadata'):
                                    layout.metadata['layout_source'] = 'adobe'
                                    layout.metadata['engine_used'] = 'adobe'  # CRITICAL: For billing
                                    layout.metadata['page_number'] = idx + 1  # CRITICAL: For billing
                                    layout.metadata['adobe_fallback_reason'] = fallback_reason
                                    layout.metadata['routing_confidence'] = adobe_confidence
                                    layout.metadata['adobe_cost_info'] = adobe_metadata
                                    layout.metadata['adobe_guardrails'] = guardrail_metadata
                                    layout.metadata['estimated_adobe_pages'] = page_count
                            
                            logger.info("=" * 80)
                            logger.info("✅ ADOBE FALLBACK: Complete - using Adobe layouts")
                            logger.info(f"Rows: {layout.metadata.get('total_rows', 'N/A')}, Columns: {layout.metadata.get('total_columns', 'N/A')}")
                            logger.info(f"Merged cells: {layout.metadata.get('merged_cells', 0)}")
                            logger.info("=" * 80)
                        else:
                            logger.warning("⚠️ Adobe fallback returned no layouts - keeping Document AI results")
                    else:
                        logger.info(f"❌ Adobe fallback requested but not available: credentials not configured")
                else:
                    logger.info(f"⏭️ Adobe fallback NOT triggered: {fallback_reason}")
                    
            except Exception as adobe_error:
                logger.warning(f"Adobe fallback check failed (non-critical): {adobe_error}")
                # Continue with Document AI results
            
            # CRITICAL FIX: Set layout_source, engine_used, and page_number metadata for DocAI (if not already set by Adobe)
            # This ensures main.py can properly calculate credits based on engine used
            for idx, layout in enumerate(unified_layouts):
                if hasattr(layout, 'metadata') and layout.metadata:
                    if 'layout_source' not in layout.metadata:
                        layout.metadata['layout_source'] = 'docai'
                    # Ensure engine_used and page_number are set for billing
                    if 'engine_used' not in layout.metadata:
                        layout.metadata['engine_used'] = layout.metadata.get('layout_source', 'docai')
                    if 'page_number' not in layout.metadata:
                        layout.metadata['page_number'] = idx + 1
                    logger.info(f"✅ Page {idx + 1}: engine={layout.metadata.get('engine_used', 'docai')}, source={layout.metadata.get('layout_source', 'docai')}")
            
            # Check if any layout has content
            has_content = any(not layout.is_empty() for layout in unified_layouts)
            
            if has_content:
                layout_strategy = "premium_layout_system"
                logger.info(f"Layout strategy: {layout_strategy}")
                logger.info(f"Generated {len(unified_layouts)} page layouts")
                
                # Extract images from document (with GCS download support)
                images = []
                try:
                    from premium_layout.image_extractor import ImageExtractor
                    from premium_layout.gcs_image_downloader import GCSImageDownloader
                    
                    image_extractor = ImageExtractor()
                    images = image_extractor.extract_images_from_document(document)
                    
                    # Download images from GCS if needed
                    gcs_downloader = GCSImageDownloader()
                    for img in images:
                        if 'gcs_uri' in img and img.get('gcs_uri') and 'image_bytes' not in img:
                            downloaded_bytes = gcs_downloader.download_image_from_gcs(img['gcs_uri'])
                            if downloaded_bytes:
                                img['image_bytes'] = downloaded_bytes
                    
                    logger.info(f"Extracted {len(images)} images from document")
                except Exception as img_error:
                    logger.warning(f"Image extraction failed (non-critical): {img_error}")
                
                # CRITICAL: Excel writing step (Step 7 of pipeline)
                # UnifiedLayout objects (from 7-step pipeline output) are written to openpyxl Workbook
                logger.info("=" * 80)
                logger.info("STEP 7: Writing UnifiedLayout to Excel (openpyxl Workbook)")
                logger.info("=" * 80)
                renderer = ExcelWordRenderer()
                excel_bytes = renderer.render_to_excel(unified_layouts, images)
                logger.info(f"STEP 7 COMPLETE: Premium layout system generated Excel: {len(excel_bytes)} bytes with {len(unified_layouts)} pages")
                logger.info("=" * 80)
                logger.info("PREMIUM PIPELINE: Complete")
                logger.info("=" * 80)
            else:
                # USER-VISIBLE BEHAVIOR: Empty layouts = blank/template document
                # Never return "object failed" - return minimal valid Excel instead
                logger.warning("Premium layout system returned empty layouts - treating as blank/template, returning minimal Excel")
                from premium_layout.unified_layout_model import UnifiedLayout, Cell, CellStyle, CellAlignment
                minimal_layout = UnifiedLayout(page_index=0)
                minimal_cell = Cell(
                    row=0,
                    column=0,
                    value="Blank or template document - minimal structure",
                    style=CellStyle(
                        bold=True,
                        alignment_horizontal=CellAlignment.CENTER,
                        background_color='FFE0E0E0'  # aRGB format (alpha + RGB)
                    )
                )
                minimal_layout.add_row([minimal_cell])
                minimal_layout.metadata['document_type'] = 'blank_template'
                minimal_layout.metadata['user_message'] = 'Blank or template document - minimal structure returned'
                
                renderer = ExcelWordRenderer()
                excel_bytes = renderer.render_to_excel([minimal_layout], [])
                logger.info(f"Generated minimal Excel for blank/template document: {len(excel_bytes)} bytes")
                
        except ImportError as import_err:
            # Premium layout system not available - this is a deployment issue
            logger.error(f"Premium layout system not available: {import_err}")
            raise Exception("Premium layout system not available. Please check deployment.")
        except Exception as premium_error:
            # USER-VISIBLE BEHAVIOR: Never return "object failed" for TYPE_A/B/D
            # Return minimal valid Excel instead
            logger.error(f"Premium layout system failed: {premium_error}")
            logger.error(f"Error details: {str(premium_error)}")
            logger.warning("Returning minimal valid Excel instead of failing")
            
            # Create minimal valid Excel
            from premium_layout.unified_layout_model import UnifiedLayout, Cell, CellStyle, CellAlignment
            from premium_layout.excel_word_renderer import ExcelWordRenderer  # CRITICAL: Re-import in except block
            
            minimal_layout = UnifiedLayout(page_index=0)
            minimal_cell = Cell(
                row=0,
                column=0,
                value="Document processing encountered an issue - minimal structure returned",
                style=CellStyle(
                    bold=True,
                    alignment_horizontal=CellAlignment.CENTER,
                    background_color='FFE0E0E0'  # aRGB format (alpha + RGB)
                )
            )
            minimal_layout.add_row([minimal_cell])
            minimal_layout.metadata['error_recovered'] = True
            minimal_layout.metadata['error_message'] = str(premium_error)
            minimal_layout.metadata['user_message'] = 'Document processing encountered an issue - minimal structure returned'
            
            try:
                renderer = ExcelWordRenderer()
                excel_bytes = renderer.render_to_excel([minimal_layout], [])
                unified_layouts = [minimal_layout]  # Set unified_layouts for return
                logger.info(f"Generated minimal Excel after error: {len(excel_bytes)} bytes")
            except Exception as render_error:
                # Last resort: create absolute minimal Excel
                logger.error(f"Even minimal Excel generation failed: {render_error}")
                # This should never happen, but if it does, we'll let the outer exception handler deal with it
                raise Exception(f"Critical error: Cannot generate Excel even with minimal structure: {str(render_error)}")
        
        # Safety check: Ensure Excel was generated
        if excel_bytes is None or len(excel_bytes) == 0:
            logger.error("Excel generation failed: No output produced")
            # USER-VISIBLE BEHAVIOR: Return minimal valid Excel instead of failing
            logger.warning("Returning minimal valid Excel instead of failing")
            from premium_layout.unified_layout_model import UnifiedLayout, Cell, CellStyle, CellAlignment
            from premium_layout.excel_word_renderer import ExcelWordRenderer  # CRITICAL: Re-import for safety check
            
            minimal_layout = UnifiedLayout(page_index=0)
            minimal_cell = Cell(
                row=0,
                column=0,
                value="Excel generation issue - minimal structure returned",
                style=CellStyle(
                    bold=True,
                    alignment_horizontal=CellAlignment.CENTER,
                    background_color='FFE0E0E0'  # aRGB format (alpha + RGB)
                )
            )
            minimal_layout.add_row([minimal_cell])
            minimal_layout.metadata['error_recovered'] = True
            minimal_layout.metadata['user_message'] = 'Excel generation issue - minimal structure returned'
            
            renderer = ExcelWordRenderer()
            excel_bytes = renderer.render_to_excel([minimal_layout], [])
            logger.info(f"Generated minimal Excel after empty output: {len(excel_bytes)} bytes")
        
        logger.info(f"Excel generation successful: {len(excel_bytes)} bytes, strategy: {layout_strategy}")
        
        # Step 6.5: Optional Excel overlay enhancement (safe hook - doesn't break pipeline if it fails)
        try:
            from excel_overlay_enhancer import enhance_excel_with_overlay
            excel_bytes = enhance_excel_with_overlay(excel_bytes, document, tables)
            logger.info("Excel overlay enhancement applied successfully")
        except ImportError:
            # Overlay module not available - use original Excel (expected in some deployments)
            logger.debug("Excel overlay enhancer not available, using original Excel")
        except Exception as overlay_error:
            # Overlay failed - use original Excel (pipeline continues normally)
            logger.warning(f"Excel overlay enhancement failed, using original Excel: {overlay_error}")
        
        # Step 7: Upload Excel to GCS and get signed URL
        logger.info("Uploading Excel to GCS...")
        base_filename = filename.replace('.pdf', '').replace('.PDF', '')
        download_url = upload_excel_to_gcs(excel_bytes, base_filename)
        logger.info(f"Excel uploaded. Download URL generated")
        
        # ENTERPRISE RESPONSE: Return unified_layouts for metadata extraction
        # main.py will extract mode, confidence, message from unified_layouts metadata
        # Ensure unified_layouts is always a list (even if empty)
        if unified_layouts is None:
            unified_layouts = []
        
        # CRITICAL: Extract pages_metadata for billing calculation (STEP-12)
        # First try to get from decision_engine (STEP-14 builds it)
        pages_metadata = []
        
        # CRITICAL FIX: decision_engine_instance is now in function scope
        if decision_engine_instance and hasattr(decision_engine_instance, 'pages_metadata'):
            pages_metadata = decision_engine_instance.pages_metadata
            logger.critical(f"✅ STEP-12: Using pages_metadata from decision_engine: {len(pages_metadata)} pages")
            for pm in pages_metadata:
                logger.critical(f"   Page {pm.get('page')}: engine={pm.get('engine')}")
        else:
            # Fallback: Extract from unified_layouts metadata
            logger.warning("⚠️ STEP-12: Extracting pages_metadata from unified_layouts (fallback)")
            if decision_engine_instance:
                logger.warning(f"   decision_engine_instance exists but no pages_metadata attr")
            else:
                logger.warning(f"   decision_engine_instance is None")
            
            for layout in unified_layouts:
                page_num = layout.metadata.get('page_number', layout.page_index + 1)
                engine = layout.metadata.get('engine_used', layout.metadata.get('layout_source', 'docai'))
                pages_metadata.append({
                    'page': page_num,
                    'engine': engine
                })
                logger.critical(f"BILLING METADATA: Page {page_num} → engine={engine}")
        
        # CRITICAL: Ensure pages_metadata is never empty if we have layouts
        if not pages_metadata and unified_layouts:
            logger.critical("⚠️ CRITICAL: pages_metadata is empty but unified_layouts exist - building from layouts")
            for idx, layout in enumerate(unified_layouts):
                page_num = layout.metadata.get('page_number', layout.page_index + 1)
                engine = layout.metadata.get('engine_used', layout.metadata.get('layout_source', 'docai'))
                pages_metadata.append({
                    'page': page_num,
                    'engine': engine
                })
                logger.critical(f"   Built metadata: Page {page_num} → engine={engine}")
        
        logger.critical(f"BILLING: Final pages_metadata count: {len(pages_metadata)} for credit calculation")
        
        return download_url, pages_processed, unified_layouts, pages_metadata
        
    except gcp_exceptions.GoogleAPIError as e:
        raise Exception(f"Document AI API error: {str(e)}")
    except Exception as e:
        logger.error(f"Error in process_pdf_to_excel_docai: {str(e)}")
        raise


def extract_tables_from_document(document) -> list:
    """
    Extract tables from Document AI document.
    Note: document type is documentai.Document but we avoid type hint to prevent import errors.
    """
    """
    Extract tables from Document AI document.
    Returns list of tables, each table is a 2D list (rows x columns).
    """
    tables = []
    
    # Log document structure for debugging
    logger.info(f"Document has {len(document.pages)} pages")
    if hasattr(document, 'text') and document.text:
        logger.info(f"Document text length: {len(document.text)} characters")
    
    # Iterate through all pages
    for page_idx, page in enumerate(document.pages):
        logger.info(f"Page {page_idx + 1}: Checking for tables...")
        
        # Check if page has tables attribute and it's not empty
        if hasattr(page, 'tables') and page.tables:
            logger.info(f"Page {page_idx + 1}: Found {len(page.tables)} table(s)")
            for table_idx, table in enumerate(page.tables):
                logger.info(f"Page {page_idx + 1}, Table {table_idx + 1}: Processing...")
                logger.info(f"  - Header rows: {len(table.header_rows) if hasattr(table, 'header_rows') and table.header_rows else 0}")
                logger.info(f"  - Body rows: {len(table.body_rows) if hasattr(table, 'body_rows') and table.body_rows else 0}")
                
                parsed_table = parse_docai_table(table, document.text if hasattr(document, 'text') else '')
                if parsed_table:
                    logger.info(f"  - Parsed table: {len(parsed_table)} rows, {len(parsed_table[0]) if parsed_table else 0} columns")
                    tables.append(parsed_table)
                else:
                    logger.warning(f"  - Table {table_idx + 1} on page {page_idx + 1} parsed as empty")
        else:
            logger.warning(f"Page {page_idx + 1}: No tables found (hasattr: {hasattr(page, 'tables')}, value: {getattr(page, 'tables', None)})")
    
    logger.info(f"Total tables extracted: {len(tables)}")
    return tables


def parse_docai_table(table, full_text: str) -> list:
    """
    Parse a Document AI table into a 2D list.
    Note: table type is documentai.Document.Page.Table but we avoid type hint to prevent import errors.
    """
    """
    Parse a single table from Document AI.
    Returns 2D list (rows x columns).
    """
    if not table.header_rows and not table.body_rows:
        return []
    
    # Combine header and body rows
    all_rows = []
    if table.header_rows:
        all_rows.extend(table.header_rows)
    if table.body_rows:
        all_rows.extend(table.body_rows)
    
    if not all_rows:
        return []
    
    # Find max columns
    max_cols = 0
    for row in all_rows:
        if row.cells:
            max_cols = max(max_cols, len(row.cells))
    
    if max_cols == 0:
        return []
    
    # Build grid
    grid = []
    for row in all_rows:
        row_data = [''] * max_cols
        if row.cells:
            for col_idx, cell in enumerate(row.cells):
                if col_idx < max_cols:
                    cell_text = extract_text_from_layout(cell.layout, full_text)
                    row_data[col_idx] = cell_text
        grid.append(row_data)
    
    return grid


def extract_text_from_layout(layout, full_text: str) -> str:
    """
    Extract text from a Document AI layout element.
    Note: layout type is documentai.Document.Page.Layout but we avoid type hint to prevent import errors.
    """
    """Extract text from a layout element using text anchor."""
    if not layout or not layout.text_anchor:
        return ''
    
    text_anchor = layout.text_anchor
    if not text_anchor.text_segments:
        return ''
    
    text_parts = []
    for segment in text_anchor.text_segments:
        start_index = int(segment.start_index) if segment.start_index else 0
        end_index = int(segment.end_index) if segment.end_index else 0
        
        if start_index < len(full_text) and end_index <= len(full_text):
            text_parts.append(full_text[start_index:end_index])
    
    return ' '.join(text_parts).strip()


def create_excel_from_docai_tables(tables: list, pdf_filename: str, document_text: str = '') -> bytes:
    """
    Create Excel workbook from Document AI tables.
    One sheet per table (or per page if multiple tables per page).
    If no tables found, attempts to extract structured data from text.
    
    Args:
        tables: List of extracted tables (2D lists)
        pdf_filename: Original PDF filename
        document_text: Full document text for fallback extraction (optional)
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    base_name = pdf_filename.replace('.pdf', '').replace('.PDF', '')
    
    # If no tables found, try to extract text and create table-like structure
    if not tables:
        logger.warning("No tables found. Attempting to extract structured data from text...")
        sheet = workbook.create_sheet(title="Extracted Content")
        
        # Add header
        sheet['A1'] = "Content"
        sheet['A1'].font = header_font
        sheet['A1'].fill = header_fill
        sheet['A1'].border = border
        sheet['A1'].alignment = center_alignment
        
        # Try to extract key-value pairs or structured data from text
        if document_text:
            logger.info(f"Processing {len(document_text)} characters of text for extraction...")
            # Split text into lines and try to identify table-like patterns
            lines = document_text.split('\n')
            logger.info(f"Split into {len(lines)} lines")
            row_num = 2
            extracted_count = 0
            
            # Look for patterns like "Key: Value" or lines with multiple spaces/tabs
            for line_idx, line in enumerate(lines[:200]):  # Increased limit to 200 lines
                original_line = line
                line = line.strip()
                if not line:
                    continue
                
                # Try to detect key-value pairs (more flexible)
                if ':' in line:
                    parts = line.split(':', 1)  # Split only on first colon
                    key = parts[0].strip()
                    value = parts[1].strip() if len(parts) > 1 else ''
                    if key and (value or len(key) < 50):  # Allow empty values if key is short
                        sheet.cell(row=row_num, column=1, value=key)
                        sheet.cell(row=row_num, column=2, value=value)
                        sheet.cell(row=row_num, column=1).border = border
                        sheet.cell(row=row_num, column=2).border = border
                        row_num += 1
                        extracted_count += 1
                        continue
                
                # Or lines with multiple spaces (potential table rows)
                if '  ' in line:  # At least one double space
                    # Split by multiple spaces (2+ spaces)
                    import re
                    parts = re.split(r'\s{2,}', line)  # Split on 2+ spaces
                    parts = [p.strip() for p in parts if p.strip()]
                    if len(parts) >= 2:
                        for col_idx, part in enumerate(parts[:15], start=1):  # Max 15 columns
                            cell = sheet.cell(row=row_num, column=col_idx, value=part)
                            cell.border = border
                        row_num += 1
                        extracted_count += 1
                        continue
                
                # Or lines with tabs
                if '\t' in line:
                    parts = [p.strip() for p in line.split('\t') if p.strip()]
                    if len(parts) >= 2:
                        for col_idx, part in enumerate(parts[:15], start=1):
                            cell = sheet.cell(row=row_num, column=col_idx, value=part)
                            cell.border = border
                        row_num += 1
                        extracted_count += 1
                        continue
                
                # Or just add as single column if no pattern detected (but only if meaningful)
                if row_num < 100 and len(line) > 3:  # Increased limit, minimum 3 chars
                    sheet.cell(row=row_num, column=1, value=line[:200])  # Increased length limit
                    sheet.cell(row=row_num, column=1).border = border
                    row_num += 1
                    extracted_count += 1
            
            logger.info(f"Extracted {extracted_count} rows of data from text (total rows: {row_num - 2})")
            
            if row_num <= 2:
                # No data extracted - add fallback message
                logger.warning("No structured data could be extracted from text")
                sheet['A2'] = "No structured tables detected in this PDF."
                sheet['A2'].font = Font(italic=True, color="666666")
                sheet['A3'] = "The PDF may contain text but no table structure."
                sheet['A3'].font = Font(italic=True, color="666666")
                sheet['A4'] = "Tip: Ensure your PDF contains clearly defined tables with rows and columns."
                sheet['A4'].font = Font(italic=True, color="666666")
            else:
                # Add column headers if we have multi-column data
                if extracted_count > 0:
                    # Auto-adjust column widths
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    cell_length = len(str(cell.value))
                                    max_length = max(max_length, cell_length)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        if adjusted_width > 0:
                            sheet.column_dimensions[column].width = adjusted_width
        else:
            # No text available
            sheet['A2'] = "No structured tables detected in this PDF."
            sheet['A2'].font = Font(italic=True, color="666666")
            sheet['A3'] = "The PDF may contain text but no table structure."
            sheet['A3'].font = Font(italic=True, color="666666")
    else:
        # Create one sheet per table
        for idx, table in enumerate(tables):
            sheet_name = f"Table {idx + 1}" if len(tables) > 1 else base_name[:31]
            sheet = workbook.create_sheet(title=sheet_name)
            
            # Write table data
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Style header row (first row)
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                    else:
                        # Apply default font to body rows (not just header)
                        cell.font = Font(size=11)  # Default font for all rows
    
    # Auto-adjust column widths
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

