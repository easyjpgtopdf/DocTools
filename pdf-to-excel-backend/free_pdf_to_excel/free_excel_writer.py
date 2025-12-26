"""
Excel Writer for FREE version.
Generates editable XLSX with basic formatting.
"""

import logging
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    HAS_OPENPYXL = True
except ImportError as e:
    HAS_OPENPYXL = False
    logger.warning(f"openpyxl not available: {e}")


def create_excel_workbook(
    grid: List[List[Dict]],
    header_rows: List[int],
    rectangles: List[Dict],
    images: List[Dict],
    row_boundaries: List[float],
    column_boundaries: List[float],
    output_path: str
) -> bool:
    """
    Create Excel workbook from grid with formatting.
    
    Args:
        grid: 2D grid of cells
        header_rows: List of row indices that are headers
        rectangles: List of rectangle objects (for background colors)
        images: List of image objects (for logos)
        row_boundaries: Row Y coordinates
        column_boundaries: Column X coordinates
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl not available")
        return False
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Page1"
        
        # Write data
        for row_idx, row in enumerate(grid):
            for col_idx, cell in enumerate(row):
                if cell.get('merged', False) and not cell.get('text'):
                    continue  # Skip merged cells
                
                cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                ws[cell_address] = cell['text']
                
                # Apply formatting
                cell_obj = ws[cell_address]
                
                # Font - Use Unicode-friendly font for better Unicode support
                font_name = cell.get('font_name', 'Arial')
                # Replace non-Unicode fonts with Unicode-friendly alternatives
                if font_name and not any(unicode_font in font_name for unicode_font in ['Arial Unicode', 'Calibri', 'Times New Roman', 'DejaVu']):
                    # Check if text contains Unicode characters
                    cell_text = cell.get('text', '')
                    has_unicode = any(ord(char) > 127 for char in str(cell_text))
                    if has_unicode:
                        # Use Unicode-friendly font
                        font_name = 'Arial Unicode MS' if 'Arial' in font_name else 'Calibri'
                
                font = Font(
                    name=font_name,
                    size=cell.get('font_size', 10),
                    bold=cell.get('is_bold', False) or (row_idx in header_rows)
                )
                cell_obj.font = font
                
                # Header background color (if rectangle aligns)
                if row_idx in header_rows:
                    # Check if rectangle aligns with this row
                    has_background = False
                    if rectangles and row_boundaries:
                        row_y = row_boundaries[row_idx] if row_idx < len(row_boundaries) else 0
                        for rect in rectangles:
                            if abs(rect['y0'] - row_y) < 20 or abs(rect['y1'] - row_y) < 20:
                                # Check if rectangle spans this column
                                if column_boundaries:
                                    col_x = column_boundaries[col_idx] if col_idx < len(column_boundaries) else 0
                                    if rect['x0'] <= col_x <= rect['x1']:
                                        has_background = True
                                        break
                    
                    if has_background:
                        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        cell_obj.fill = fill
                
                # Alignment
                cell_obj.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Basic borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell_obj.border = thin_border
        
        # Apply merged cells
        merge_ranges = []
        for row_idx, row in enumerate(grid):
            for col_idx, cell in enumerate(row):
                if cell.get('merged', False) and not cell.get('text'):
                    continue  # Skip empty merged cells
                
                # Check for horizontal merge (merge_right)
                if cell.get('merge_right', False):
                    # Find how many cells to merge horizontally
                    merge_end_col = col_idx
                    for c in range(col_idx + 1, len(row)):
                        if row[c].get('merge_left', False) and not row[c].get('text'):
                            merge_end_col = c
                        else:
                            break
                    
                    if merge_end_col > col_idx:
                        start_cell = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        end_cell = f"{get_column_letter(merge_end_col + 1)}{row_idx + 1}"
                        merge_ranges.append((start_cell, end_cell))
                
                # Check for vertical merge (merge_down)
                if cell.get('merge_down', False):
                    # Find how many cells to merge vertically
                    merge_end_row = row_idx
                    for r in range(row_idx + 1, len(grid)):
                        if col_idx < len(grid[r]) and grid[r][col_idx].get('merge_up', False) and not grid[r][col_idx].get('text'):
                            merge_end_row = r
                        else:
                            break
                    
                    if merge_end_row > row_idx:
                        start_cell = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        end_cell = f"{get_column_letter(col_idx + 1)}{merge_end_row + 1}"
                        merge_ranges.append((start_cell, end_cell))
        
        # Apply merges to worksheet
        for start_cell, end_cell in merge_ranges:
            try:
                ws.merge_cells(f"{start_cell}:{end_cell}")
                logger.debug(f"Merged cells: {start_cell}:{end_cell}")
            except Exception as merge_error:
                logger.warning(f"Could not merge {start_cell}:{end_cell}: {merge_error}")
        
        # Auto-size columns
        if grid and len(grid) > 0 and len(grid[0]) > 0:
            for col_idx in range(len(grid[0])):
                col_letter = get_column_letter(col_idx + 1)
                max_length = 0
                for row in grid:
                    if col_idx < len(row):
                        cell_text = str(row[col_idx].get('text', ''))
                        max_length = max(max_length, len(cell_text))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Place images (logos) as floating objects in Excel
        logger.info(f"Attempting to insert {len(images)} images into Excel")
        images_inserted = 0
        for img in images:
            try:
                # Check if image data is available
                image_data = img.get('image_data')
                if not image_data:
                    logger.warning(f"Image at ({img['x0']}, {img['y0']}) has no image data, skipping")
                    continue
                
                # Find nearest cell for anchor
                if column_boundaries and row_boundaries:
                    col_idx = 0
                    min_col_dist = float('inf')
                    for i, col_x in enumerate(column_boundaries):
                        dist = abs(img['x0'] - col_x)
                        if dist < min_col_dist:
                            min_col_dist = dist
                            col_idx = i
                    
                    row_idx = 0
                    min_row_dist = float('inf')
                    for i, row_y in enumerate(row_boundaries):
                        dist = abs(img['y0'] - row_y)
                        if dist < min_row_dist:
                            min_row_dist = dist
                            row_idx = i
                    
                    # Create image from bytes
                    from io import BytesIO
                    
                    # Validate image data
                    if not isinstance(image_data, bytes):
                        logger.warning(f"Image data is not bytes, converting...")
                        try:
                            image_data = bytes(image_data)
                        except Exception as conv_e:
                            logger.error(f"Cannot convert image data to bytes: {conv_e}")
                            continue
                    
                    # Check if image data is valid
                    if len(image_data) < 10:  # Too small to be a valid image
                        logger.warning(f"Image data too small ({len(image_data)} bytes), skipping")
                        continue
                    
                    image_stream = BytesIO(image_data)
                    
                    # Create openpyxl Image object
                    try:
                        img_obj = OpenpyxlImage(image_stream)
                    except Exception as img_create_e:
                        logger.error(f"Failed to create openpyxl Image object: {img_create_e}")
                        # Try to fix image data - maybe it needs decoding
                        try:
                            # Some PDF images are compressed, try to decompress
                            import zlib
                            try:
                                decompressed = zlib.decompress(image_data)
                                image_stream = BytesIO(decompressed)
                                img_obj = OpenpyxlImage(image_stream)
                                logger.info(f"✅ Image decompressed successfully")
                            except:
                                # If decompression fails, try as-is with different format
                                image_stream = BytesIO(image_data)
                                img_obj = OpenpyxlImage(image_stream)
                        except Exception as fix_e:
                            logger.error(f"Could not fix image data: {fix_e}")
                            continue
                    
                    # Scale image to fit (max 150x150 pixels for better visibility)
                    max_cell_size = 150
                    img_width = img.get('width', 100)
                    img_height = img.get('height', 100)
                    
                    # Calculate scale to fit within max_cell_size
                    # Preserve aspect ratio
                    if img_width > 0 and img_height > 0:
                        scale = min(max_cell_size / max(img_width, img_height), 1.0)
                        img_obj.width = int(img_width * scale)
                        img_obj.height = int(img_height * scale)
                    else:
                        # Default size if dimensions unknown
                        img_obj.width = 100
                        img_obj.height = 100
                    
                    # Anchor image to cell (top-left corner)
                    cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                    img_obj.anchor = cell_address
                    
                    # Add image to worksheet
                    ws.add_image(img_obj)
                    images_inserted += 1
                    
                    logger.info(f"✅ Image {images_inserted} inserted at cell {cell_address} (size: {img_obj.width}x{img_obj.height}, format: {img.get('image_format', 'unknown')})")
                    
            except Exception as e:
                logger.error(f"Error inserting image at ({img.get('x0', 0)}, {img.get('y0', 0)}): {e}")
                import traceback
                logger.error(traceback.format_exc())
                # Fallback: add comment if image insertion fails
                try:
                    if column_boundaries and row_boundaries:
                        col_idx = 0
                        min_col_dist = float('inf')
                        for i, col_x in enumerate(column_boundaries):
                            dist = abs(img['x0'] - col_x)
                            if dist < min_col_dist:
                                min_col_dist = dist
                                col_idx = i
                        
                        row_idx = 0
                        min_row_dist = float('inf')
                        for i, row_y in enumerate(row_boundaries):
                            dist = abs(img['y0'] - row_y)
                            if dist < min_row_dist:
                                min_row_dist = dist
                                row_idx = i
                        
                        if row_idx < len(grid) and col_idx < len(grid[row_idx]):
                            cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                            cell_obj = ws[cell_address]
                            if not cell_obj.comment:
                                from openpyxl.comments import Comment
                                cell_obj.comment = Comment("Image/Logo detected (insertion failed)", "PDF Converter")
                                fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                                cell_obj.fill = fill
                except Exception as fallback_error:
                    logger.error(f"Fallback comment also failed: {fallback_error}")
        
        logger.info(f"Excel creation complete: {images_inserted}/{len(images)} images inserted successfully")
        wb.save(output_path)
        return True
    
    except Exception as e:
        logger.error(f"Error creating Excel workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def create_excel_workbook_multi_page(
    all_pages_data: List[Dict],
    output_path: str
) -> bool:
    """
    Create Excel workbook with multiple sheets (one per page).
    
    Args:
        all_pages_data: List of page data dictionaries, each containing:
            - page_num: Page index (0-based)
            - grid: 2D grid of cells
            - header_rows: List of row indices that are headers
            - rectangles: List of rectangle objects
            - images: List of image objects
            - row_boundaries: Row Y coordinates
            - column_boundaries: Column X coordinates
            - grid_confidence: Confidence score
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl not available")
        return False
    
    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if wb.worksheets:
            wb.remove(wb.worksheets[0])
        
        total_images_inserted = 0
        
        # Process each page
        for page_data in all_pages_data:
            page_num = page_data['page_num']
            grid = page_data['grid']
            header_rows = page_data['header_rows']
            rectangles = page_data['rectangles']
            images = page_data['images']
            row_boundaries = page_data['row_boundaries']
            column_boundaries = page_data['column_boundaries']
            
            # Create sheet for this page
            ws = wb.create_sheet(title=f"Page{page_num + 1}")
            
            # Write data
            for row_idx, row in enumerate(grid):
                for col_idx, cell in enumerate(row):
                    if cell.get('merged', False) and not cell.get('text'):
                        continue  # Skip merged cells (will be merged later)
                    
                    cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                    ws[cell_address] = cell['text']
                    
                    # Apply formatting
                    cell_obj = ws[cell_address]
                    
                    # Font - Use Unicode-friendly font for better Unicode support
                    font_name = cell.get('font_name', 'Arial')
                    if font_name and not any(unicode_font in font_name for unicode_font in ['Arial Unicode', 'Calibri', 'Times New Roman', 'DejaVu']):
                        cell_text = cell.get('text', '')
                        has_unicode = any(ord(char) > 127 for char in str(cell_text))
                        if has_unicode:
                            font_name = 'Arial Unicode MS' if 'Arial' in font_name else 'Calibri'
                    
                    font = Font(
                        name=font_name,
                        size=cell.get('font_size', 10),
                        bold=cell.get('is_bold', False) or (row_idx in header_rows)
                    )
                    cell_obj.font = font
                    
                    # Header background color
                    if row_idx in header_rows:
                        has_background = False
                        if rectangles and row_boundaries:
                            row_y = row_boundaries[row_idx] if row_idx < len(row_boundaries) else 0
                            for rect in rectangles:
                                if abs(rect['y0'] - row_y) < 20 or abs(rect['y1'] - row_y) < 20:
                                    if column_boundaries:
                                        col_x = column_boundaries[col_idx] if col_idx < len(column_boundaries) else 0
                                        if rect['x0'] <= col_x <= rect['x1']:
                                            has_background = True
                                            break
                        
                        if has_background:
                            fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell_obj.fill = fill
                    
                    # Alignment
                    cell_obj.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    
                    # Basic borders
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell_obj.border = thin_border
            
            # Apply merged cells for this page
            merge_ranges = []
            for row_idx, row in enumerate(grid):
                for col_idx, cell in enumerate(row):
                    if cell.get('merged', False) and not cell.get('text'):
                        continue
                    
                    # Check for horizontal merge
                    if cell.get('merge_right', False):
                        merge_end_col = col_idx
                        for c in range(col_idx + 1, len(row)):
                            if row[c].get('merge_left', False) and not row[c].get('text'):
                                merge_end_col = c
                            else:
                                break
                        
                        if merge_end_col > col_idx:
                            start_cell = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                            end_cell = f"{get_column_letter(merge_end_col + 1)}{row_idx + 1}"
                            merge_ranges.append((start_cell, end_cell))
                    
                    # Check for vertical merge
                    if cell.get('merge_down', False):
                        merge_end_row = row_idx
                        for r in range(row_idx + 1, len(grid)):
                            if col_idx < len(grid[r]) and grid[r][col_idx].get('merge_up', False) and not grid[r][col_idx].get('text'):
                                merge_end_row = r
                            else:
                                break
                        
                        if merge_end_row > row_idx:
                            start_cell = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                            end_cell = f"{get_column_letter(col_idx + 1)}{merge_end_row + 1}"
                            merge_ranges.append((start_cell, end_cell))
            
            # Apply merges to worksheet
            for start_cell, end_cell in merge_ranges:
                try:
                    ws.merge_cells(f"{start_cell}:{end_cell}")
                    logger.debug(f"Page {page_num + 1}: Merged cells {start_cell}:{end_cell}")
                except Exception as merge_error:
                    logger.warning(f"Page {page_num + 1}: Could not merge {start_cell}:{end_cell}: {merge_error}")
            
            # Auto-size columns
            if grid and len(grid) > 0 and len(grid[0]) > 0:
                for col_idx in range(len(grid[0])):
                    col_letter = get_column_letter(col_idx + 1)
                    max_length = 0
                    for row in grid:
                        if col_idx < len(row):
                            cell_text = str(row[col_idx].get('text', ''))
                            max_length = max(max_length, len(cell_text))
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            # Place images (logos) as floating objects in Excel
            logger.info(f"Page {page_num + 1}: Attempting to insert {len(images)} images into Excel")
            images_inserted = 0
            for img in images:
                try:
                    image_data = img.get('image_data')
                    if not image_data:
                        logger.warning(f"Image at ({img['x0']}, {img['y0']}) has no image data, skipping")
                        continue
                    
                    # Find nearest cell for anchor
                    if column_boundaries and row_boundaries:
                        col_idx = 0
                        min_col_dist = float('inf')
                        for i, col_x in enumerate(column_boundaries):
                            dist = abs(img['x0'] - col_x)
                            if dist < min_col_dist:
                                min_col_dist = dist
                                col_idx = i
                        
                        row_idx = 0
                        min_row_dist = float('inf')
                        for i, row_y in enumerate(row_boundaries):
                            dist = abs(img['y0'] - row_y)
                            if dist < min_row_dist:
                                min_row_dist = dist
                                row_idx = i
                        
                        # Create image from bytes
                        from io import BytesIO
                        
                        if not isinstance(image_data, bytes):
                            try:
                                image_data = bytes(image_data)
                            except Exception as conv_e:
                                logger.error(f"Cannot convert image data to bytes: {conv_e}")
                                continue
                        
                        if len(image_data) < 10:
                            logger.warning(f"Image data too small ({len(image_data)} bytes), skipping")
                            continue
                        
                        image_stream = BytesIO(image_data)
                        
                        try:
                            img_obj = OpenpyxlImage(image_stream)
                        except Exception as img_create_e:
                            logger.error(f"Failed to create openpyxl Image object: {img_create_e}")
                            try:
                                import zlib
                                try:
                                    decompressed = zlib.decompress(image_data)
                                    image_stream = BytesIO(decompressed)
                                    img_obj = OpenpyxlImage(image_stream)
                                    logger.info(f"✅ Image decompressed successfully")
                                except:
                                    image_stream = BytesIO(image_data)
                                    img_obj = OpenpyxlImage(image_stream)
                            except Exception as fix_e:
                                logger.error(f"Could not fix image data: {fix_e}")
                                continue
                        
                        # Scale image to fit
                        max_cell_size = 150
                        img_width = img.get('width', 100)
                        img_height = img.get('height', 100)
                        
                        if img_width > 0 and img_height > 0:
                            scale = min(max_cell_size / max(img_width, img_height), 1.0)
                            img_obj.width = int(img_width * scale)
                            img_obj.height = int(img_height * scale)
                        else:
                            img_obj.width = 100
                            img_obj.height = 100
                        
                        # Anchor image to cell
                        cell_address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        img_obj.anchor = cell_address
                        
                        # Add image to worksheet
                        ws.add_image(img_obj)
                        images_inserted += 1
                        total_images_inserted += 1
                        
                        logger.info(f"✅ Page {page_num + 1}: Image {images_inserted} inserted at cell {cell_address}")
                        
                except Exception as e:
                    logger.error(f"Error inserting image on page {page_num + 1}: {e}")
                    continue
            
            logger.info(f"Page {page_num + 1}: {images_inserted}/{len(images)} images inserted successfully")
        
        logger.info(f"Excel creation complete: {len(all_pages_data)} pages, {total_images_inserted} total images inserted")
        wb.save(output_path)
        return True
    
    except Exception as e:
        logger.error(f"Error creating multi-page Excel workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

