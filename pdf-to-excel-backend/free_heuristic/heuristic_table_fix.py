"""
Heuristic Table Fix for FREE PDF to Excel Pipeline.
Fixes collapsed rows, misaligned label-value pairs, missing row breaks.

Works ONLY on Excel cell content (no PDF parsing).
CPU-only, O(n log n) complexity.
"""

import logging
from typing import List, Dict, Tuple, Optional
import copy
import re

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not available for heuristic table fix")


def fix_table_layout(excel_path: str, doc_type: str = None) -> bool:
    """
    Apply heuristic fixes to Excel table layout.
    
    Args:
        excel_path: Path to Excel file (modified in-place)
        doc_type: Document type (for type-specific fixes)
        
    Returns:
        True if fixes were applied, False otherwise
    """
    if not HAS_OPENPYXL:
        return False
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Read all cell values into memory (O(n))
        rows_data = []
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                row_data.append({
                    'value': str(value) if value is not None else '',
                    'row': row_idx,
                    'col': col_idx,
                    'cell': cell
                })
            rows_data.append(row_data)
        
        # Apply document-type-specific fixes
        if doc_type:
            logger.info(f"Applying document-type-specific fixes for: {doc_type}")
            fixed_rows = apply_document_type_fixes(rows_data, doc_type)
        else:
            # Fallback: Check if this is a single-column layout (form document)
            single_col_layout = detect_single_column_layout(rows_data)
            
            if single_col_layout:
                logger.info("Detected single-column layout (form document), applying form-specific fixes")
                fixed_rows = fix_form_layout(rows_data)
            else:
                # Apply standard heuristic fixes
                fixed_rows = apply_heuristic_fixes(rows_data)
        
        if fixed_rows == rows_data:
            # No changes made
            return False
        
        # Write fixed data back to Excel
        # Clear existing content
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col_idx).value = None
        
        # Write fixed content
        max_fixed_col = 1
        for row_data in fixed_rows:
            for cell_data in row_data:
                if cell_data['value']:
                    col = cell_data['col']
                    if col > max_fixed_col:
                        max_fixed_col = col
                    ws.cell(row=cell_data['row'], column=col).value = cell_data['value']
        
        wb.save(excel_path)
        logger.info("Heuristic table fixes applied successfully")
        return True
        
    except Exception as e:
        logger.error(f"Error applying heuristic table fixes: {e}")
        return False


def apply_document_type_fixes(rows_data: List[List[Dict]], doc_type: str) -> List[List[Dict]]:
    """
    Apply document-type-specific fixes.
    
    Args:
        rows_data: List of rows (each row is list of cell dicts)
        doc_type: Document type
        
    Returns:
        Fixed rows data
    """
    if doc_type == 'invoice' or doc_type == 'bank_statement':
        # Trust LibreOffice table, only fix:
        # - merged rows
        # - empty row gaps
        # - misaligned numeric columns
        return fix_invoice_bank_layout(rows_data)
    
    elif doc_type == 'bill_or_receipt':
        # Apply label-value inference
        # Left column = label, Right column = value
        # Preserve original reading order
        return fix_bill_receipt_layout(rows_data)
    
    elif doc_type == 'resume':
        # Convert sections into rows: Name | Value, Section | Content
        # Bullet points become separate rows
        return fix_resume_layout(rows_data)
    
    elif doc_type == 'certificate' or doc_type == 'id_card':
        # Force 2-column layout: Field | Value
        # Multiple values split into new rows
        # Never guess missing data
        return fix_certificate_id_layout(rows_data)
    
    elif doc_type == 'letter':
        # Paragraphs become rows
        # Keep single-column layout
        # Maintain reading order
        return fix_letter_layout(rows_data)
    
    elif doc_type == 'generic_form':
        # Apply form layout fixes
        return fix_form_layout(rows_data)
    
    else:
        # Unknown type - apply generic fixes
        return apply_heuristic_fixes(rows_data)


def fix_invoice_bank_layout(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix invoice/bank statement layout.
    Trust LibreOffice table, only fix:
    - merged rows
    - empty row gaps
    - misaligned numeric columns
    """
    fixed_rows = []
    
    # Remove empty rows
    for row_data in rows_data:
        if any(cell.get('value', '').strip() for cell in row_data):
            fixed_rows.append(row_data)
    
    # Fix merged rows (rows with duplicate content)
    i = 0
    while i < len(fixed_rows):
        row = fixed_rows[i]
        
        # Check if next row is duplicate (merged row)
        if i + 1 < len(fixed_rows):
            next_row = fixed_rows[i + 1]
            if rows_are_duplicate(row, next_row):
                # Skip duplicate row
                i += 1
                continue
        
        fixed_rows[i] = row
        i += 1
    
    # Fix misaligned numeric columns (align numbers to right)
    # This is handled by Excel formatting, so we just ensure structure is correct
    return fixed_rows


def fix_bill_receipt_layout(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix bill/receipt layout.
    Apply label-value inference:
    - Left column = label
    - Right column = value
    - Preserve original reading order
    
    Handles cases where:
    1. Label:Value in single cell -> Split into 2 columns
    2. Labels in Column A, values missing -> Look in adjacent rows
    3. Labels in Column A, values in Column B but misaligned -> Reconstruct pairs
    """
    fixed_rows = []
    i = 0
    
    while i < len(rows_data):
        row_data = rows_data[i]
        # Remove empty rows
        non_empty = [cell for cell in row_data if cell.get('value', '').strip()]
        if not non_empty:
            i += 1
            continue
        
        # Case 1: Single column with "Label: Value" pattern
        if len(non_empty) == 1:
            cell = non_empty[0]
            value = cell['value'].strip()
            
            # Check for label:value pattern
            if ':' in value:
                parts = value.split(':', 1)
                if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                    label = parts[0].strip()
                    val = parts[1].strip()
                    
                    # Create 2-column row
                    new_row = copy.deepcopy(row_data)
                    new_row[0]['value'] = label
                    if len(new_row) > 1:
                        new_row[1]['value'] = val
                    else:
                        new_row.append({
                            'value': val,
                            'row': cell['row'],
                            'col': 2,
                            'cell': None
                        })
                    fixed_rows.append(new_row)
                    i += 1
                    continue
        
        # Case 2: Label in Column A (col=1), but Column B is empty or has wrong value
        # Check if first column has a label-like text (short, no numbers, ends with colon or is a field name)
        first_col_cell = row_data[0] if row_data else None
        if first_col_cell and first_col_cell.get('value', '').strip():
            label_text = first_col_cell['value'].strip()
            is_label = (
                len(label_text) < 50 and  # Short text
                (':' in label_text or  # Has colon
                 label_text.lower() in ['gst registration number', 'website url', 'agent name & address', 
                                       'agent id', 'transaction date', 'total amount', 'biller platform fee',
                                       'customer convenience fee', 'bill amount', 'bill date', 
                                       'payment channel', 'payment status', 'name of the customer',
                                       'biller name', 'biller id', 'b-connect txn id', 'approval ref no',
                                       'consumer number/ivrs', 'mobile number', 'payment mode']) or
                any(keyword in label_text.lower() for keyword in ['name', 'date', 'number', 'id', 'amount', 'fee', 'status', 'channel', 'mode'])
            )
            
            if is_label:
                # Check Column B (index 1)
                second_col_cell = row_data[1] if len(row_data) > 1 else None
                second_col_value = second_col_cell.get('value', '').strip() if second_col_cell else ''
                
                # If Column B is empty, look for value in next row
                if not second_col_value:
                    # Look ahead in next 2 rows for a value
                    value_found = None
                    for j in range(i + 1, min(i + 3, len(rows_data))):
                        next_row = rows_data[j]
                        # Check if next row's first column has a value (might be the value for this label)
                        if next_row and len(next_row) > 0:
                            next_first = next_row[0].get('value', '').strip()
                            # If next row's first column looks like a value (not a label), use it
                            if next_first and len(next_first) > 0:
                                # Check if it's not a label (no colon, not too short, has content)
                                if ':' not in next_first and len(next_first) > 2:
                                    value_found = next_first
                                    # Mark next row as consumed (skip it)
                                    i = j + 1
                                    break
                                # Also check next row's second column
                                if len(next_row) > 1:
                                    next_second = next_row[1].get('value', '').strip()
                                    if next_second and len(next_second) > 2:
                                        value_found = next_second
                                        i = j + 1
                                        break
                    
                    # Create 2-column row with label and found value
                    new_row = copy.deepcopy(row_data)
                    new_row[0]['value'] = label_text
                    if len(new_row) > 1:
                        new_row[1]['value'] = value_found if value_found else ''
                    else:
                        new_row.append({
                            'value': value_found if value_found else '',
                            'row': first_col_cell['row'],
                            'col': 2,
                            'cell': None
                        })
                    fixed_rows.append(new_row)
                    continue
        
        # Case 3: Already 2+ columns, ensure left=label, right=value
        if len(non_empty) >= 2:
            # First non-empty = label, last non-empty = value
            new_row = copy.deepcopy(row_data)
            new_row[non_empty[0]['col'] - 1]['value'] = non_empty[0]['value']
            new_row[non_empty[-1]['col'] - 1]['value'] = non_empty[-1]['value']
            fixed_rows.append(new_row)
        else:
            fixed_rows.append(row_data)
        
        i += 1
    
    return fixed_rows


def fix_resume_layout(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix resume layout.
    Convert sections into rows:
    - Name | Value
    - Section | Content
    - Bullet points become separate rows
    """
    fixed_rows = []
    
    for row_data in rows_data:
        non_empty = [cell for cell in row_data if cell.get('value', '').strip()]
        if not non_empty:
            continue
        
        # Check for bullet points (•, -, *, etc.)
        for cell in non_empty:
            value = cell['value'].strip()
            if value.startswith(('•', '-', '*', '·')) or value.startswith(('1.', '2.', '3.', '4.', '5.')):
                # Bullet point - create separate row
                new_row = copy.deepcopy(row_data)
                new_row[cell['col'] - 1]['value'] = value
                # Clear other cells
                for c in new_row:
                    if c['col'] != cell['col']:
                        c['value'] = ''
                fixed_rows.append(new_row)
                continue
        
        # Check for section headers (all caps, short text)
        if len(non_empty) == 1:
            value = non_empty[0]['value'].strip()
            # If all caps and short, it's likely a section header
            if value.isupper() and len(value.split()) <= 5:
                new_row = copy.deepcopy(row_data)
                new_row[0]['value'] = value
                if len(new_row) > 1:
                    new_row[1]['value'] = ''  # Section content will be in next rows
                fixed_rows.append(new_row)
                continue
        
        # Regular row - ensure 2-column format (Name | Value or Section | Content)
        new_row = copy.deepcopy(row_data)
        if len(non_empty) >= 2:
            new_row[0]['value'] = non_empty[0]['value']
            new_row[1]['value'] = ' '.join([c['value'] for c in non_empty[1:]])
        else:
            new_row[0]['value'] = non_empty[0]['value']
            if len(new_row) > 1:
                new_row[1]['value'] = ''
        fixed_rows.append(new_row)
    
    return fixed_rows


def fix_certificate_id_layout(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix certificate/ID card layout.
    Force 2-column layout: Field | Value
    Multiple values split into new rows
    Never guess missing data
    """
    fixed_rows = []
    
    for row_data in rows_data:
        non_empty = [cell for cell in row_data if cell.get('value', '').strip()]
        if not non_empty:
            continue
        
        # Force 2-column layout
        new_row = copy.deepcopy(row_data)
        
        # Ensure we have at least 2 columns
        while len(new_row) < 2:
            new_row.append({
                'value': '',
                'row': row_data[0]['row'],
                'col': len(new_row) + 1,
                'cell': None
            })
        
        # First non-empty = Field, rest = Value
        if len(non_empty) >= 1:
            new_row[0]['value'] = non_empty[0]['value']
            if len(non_empty) > 1:
                # Multiple values - split into separate rows
                for val_cell in non_empty[1:]:
                    val_row = copy.deepcopy(new_row)
                    val_row[0]['value'] = non_empty[0]['value']  # Keep field name
                    val_row[1]['value'] = val_cell['value']
                    fixed_rows.append(val_row)
            else:
                new_row[1]['value'] = ''
                fixed_rows.append(new_row)
        else:
            fixed_rows.append(new_row)
    
    return fixed_rows


def fix_letter_layout(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix letter layout.
    Paragraphs become rows
    Keep single-column layout
    Maintain reading order
    """
    fixed_rows = []
    
    current_paragraph = []
    
    for row_data in rows_data:
        non_empty = [cell for cell in row_data if cell.get('value', '').strip()]
        if not non_empty:
            # Empty row = paragraph break
            if current_paragraph:
                # Combine paragraph into single row
                para_text = ' '.join([cell['value'] for cell in current_paragraph])
                new_row = copy.deepcopy(row_data)
                new_row[0]['value'] = para_text
                fixed_rows.append(new_row)
                current_paragraph = []
            continue
        
        # Add to current paragraph
        current_paragraph.extend(non_empty)
    
    # Add last paragraph if exists
    if current_paragraph:
        para_text = ' '.join([cell['value'] for cell in current_paragraph])
        new_row = copy.deepcopy(rows_data[-1] if rows_data else [])
        if not new_row:
            new_row = [{'value': '', 'row': 1, 'col': 1, 'cell': None}]
        new_row[0]['value'] = para_text
        fixed_rows.append(new_row)
    
    return fixed_rows if fixed_rows else rows_data


def rows_are_duplicate(row1: List[Dict], row2: List[Dict]) -> bool:
    """Check if two rows are duplicates (merged rows)."""
    values1 = [cell.get('value', '').strip() for cell in row1]
    values2 = [cell.get('value', '').strip() for cell in row2]
    return values1 == values2 and any(values1)


def apply_heuristic_fixes(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Apply heuristic fixes to row data.
    
    Args:
        rows_data: List of rows, each row is a list of cell dictionaries
        
    Returns:
        Fixed rows data (new list, original unchanged)
    """
    if not rows_data:
        return rows_data
    
    # Create a copy to avoid modifying original
    fixed_rows = copy.deepcopy(rows_data)
    
    # Fix 1: Detect and split collapsed rows (rows with too much content in one cell)
    fixed_rows = fix_collapsed_rows(fixed_rows)
    
    # Fix 2: Detect and align label-value pairs (patterns like "Label: Value")
    fixed_rows = fix_label_value_pairs(fixed_rows)
    
    # Fix 3: Add missing row breaks (detect multi-line content that should be separate rows)
    fixed_rows = fix_missing_row_breaks(fixed_rows)
    
    return fixed_rows


def fix_collapsed_rows(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix collapsed rows where multiple logical rows are in one physical row.
    
    Strategy: If a cell has multiple sentences or line breaks, split it.
    """
    fixed_rows = []
    
    for row_data in rows_data:
        # Check if any cell has multiple sentences (indicated by periods or line breaks)
        needs_split = False
        for cell_data in row_data:
            value = cell_data['value']
            if value and (value.count('.') > 2 or '\n' in value or '  ' in value):
                needs_split = True
                break
        
        if needs_split:
            # Try to split the row
            split_rows = split_row_by_content(row_data)
            fixed_rows.extend(split_rows)
        else:
            fixed_rows.append(row_data)
    
    return fixed_rows


def split_row_by_content(row_data: List[Dict]) -> List[List[Dict]]:
    """
    Split a row into multiple rows based on content patterns.
    
    Returns:
        List of row data (each is a list of cell dictionaries)
    """
    if not row_data:
        return [row_data]
    
    # Find the cell with the most content
    max_len = 0
    max_idx = 0
    for idx, cell_data in enumerate(row_data):
        if len(cell_data['value']) > max_len:
            max_len = len(cell_data['value'])
            max_idx = idx
    
    # If the max cell has line breaks, split by line breaks
    max_cell = row_data[max_idx]
    value = max_cell['value']
    
    if '\n' in value:
        lines = value.split('\n')
        split_rows = []
        for line_idx, line in enumerate(lines):
            if line.strip():
                new_row = copy.deepcopy(row_data)
                new_row[max_idx]['value'] = line.strip()
                # Clear other cells in this row (keep only the split content)
                for cell_idx, cell_data in enumerate(new_row):
                    if cell_idx != max_idx:
                        cell_data['value'] = ''
                split_rows.append(new_row)
        return split_rows if split_rows else [row_data]
    
    # If the max cell has multiple sentences, try to split by periods
    if value.count('.') > 2:
        sentences = [s.strip() + '.' for s in value.split('.') if s.strip()]
        if len(sentences) > 1:
            split_rows = []
            for sent in sentences:
                new_row = copy.deepcopy(row_data)
                new_row[max_idx]['value'] = sent
                for cell_idx, cell_data in enumerate(new_row):
                    if cell_idx != max_idx:
                        cell_data['value'] = ''
                split_rows.append(new_row)
            return split_rows if split_rows else [row_data]
    
    return [row_data]


def fix_label_value_pairs(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix misaligned label-value pairs.
    
    Strategy: Detect "Label: Value" patterns and ensure they're in separate columns.
    """
    fixed_rows = []
    
    for row_data in rows_data:
        # Check if any cell contains ":" pattern (label: value)
        has_colon_pattern = False
        colon_cell_idx = -1
        
        for idx, cell_data in enumerate(row_data):
            value = cell_data['value']
            if ':' in value and len(value.split(':')) == 2:
                parts = value.split(':')
                if len(parts[0].strip()) > 0 and len(parts[1].strip()) > 0:
                    has_colon_pattern = True
                    colon_cell_idx = idx
                    break
        
        if has_colon_pattern and colon_cell_idx >= 0:
            # Split label and value into separate columns
            cell_data = row_data[colon_cell_idx]
            value = cell_data['value']
            label, value_part = value.split(':', 1)
            label = label.strip()
            value_part = value_part.strip()
            
            # Create new row with label and value in adjacent columns
            new_row = copy.deepcopy(row_data)
            new_row[colon_cell_idx]['value'] = label
            
            # Put value in next column if available
            if colon_cell_idx + 1 < len(new_row):
                new_row[colon_cell_idx + 1]['value'] = value_part
            else:
                # Add a new column (extend row)
                new_cell = {
                    'value': value_part,
                    'row': cell_data['row'],
                    'col': cell_data['col'] + 1,
                    'cell': None  # Will be created when writing back
                }
                new_row.append(new_cell)
            
            fixed_rows.append(new_row)
        else:
            fixed_rows.append(row_data)
    
    return fixed_rows


def fix_missing_row_breaks(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix missing row breaks where content should be on separate rows.
    
    Strategy: Detect cells with multiple distinct pieces of information.
    """
    fixed_rows = []
    
    for row_data in rows_data:
        # Check for cells with multiple distinct values (separated by common delimiters)
        needs_break = False
        
        for cell_data in row_data:
            value = cell_data['value']
            if not value:
                continue
            
            # Check for multiple distinct pieces (separated by |, ;, or multiple spaces)
            if '|' in value:
                parts = [p.strip() for p in value.split('|') if p.strip()]
                if len(parts) > 1:
                    needs_break = True
                    break
            elif ';' in value:
                parts = [p.strip() for p in value.split(';') if p.strip()]
                if len(parts) > 1:
                    needs_break = True
                    break
        
        if needs_break:
            # Split into multiple rows
            split_rows = split_row_by_delimiters(row_data)
            fixed_rows.extend(split_rows)
        else:
            fixed_rows.append(row_data)
    
    return fixed_rows


def split_row_by_delimiters(row_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Split a row by delimiter-separated values.
    """
    if not row_data:
        return [row_data]
    
    # Find cell with delimiter
    delimiter = None
    delimiter_idx = -1
    
    for idx, cell_data in enumerate(row_data):
        value = cell_data['value']
        if '|' in value:
            delimiter = '|'
            delimiter_idx = idx
            break
        elif ';' in value:
            delimiter = ';'
            delimiter_idx = idx
            break
    
    if delimiter_idx < 0:
        return [row_data]
    
    # Split the cell value
    cell_data = row_data[delimiter_idx]
    value = cell_data['value']
    parts = [p.strip() for p in value.split(delimiter) if p.strip()]
    
    if len(parts) <= 1:
        return [row_data]
    
    # Create separate rows for each part
    split_rows = []
    for part in parts:
        new_row = copy.deepcopy(row_data)
        new_row[delimiter_idx]['value'] = part
        # Clear other cells
        for cell_idx, cell_data in enumerate(new_row):
            if cell_idx != delimiter_idx:
                cell_data['value'] = ''
        split_rows.append(new_row)
    
    return split_rows if split_rows else [row_data]


def detect_single_column_layout(rows_data: List[List[Dict]]) -> bool:
    """
    Detect if the layout is a single-column form (labels in one column).
    
    Returns:
        True if single-column layout detected
    """
    if not rows_data:
        return False
    
    # Count rows with content only in first column
    single_col_rows = 0
    multi_col_rows = 0
    
    for row_data in rows_data:
        if not row_data:
            continue
        
        # Count non-empty cells
        non_empty = [cell for cell in row_data if cell.get('value', '').strip()]
        
        if not non_empty:
            continue
        
        # If only first column has content, it's single-column
        if len(non_empty) == 1 and non_empty[0]['col'] == 1:
            single_col_rows += 1
        elif len(non_empty) > 1:
            multi_col_rows += 1
    
    # If more than 70% of rows are single-column, it's a form layout
    total_rows = single_col_rows + multi_col_rows
    if total_rows == 0:
        return False
    
    ratio = single_col_rows / total_rows
    return ratio >= 0.7


def fix_form_layout(rows_data: List[List[Dict]]) -> List[List[Dict]]:
    """
    Fix form layout (single-column labels) by:
    1. Removing empty rows
    2. Creating 2-column layout (Label | Value)
    3. Grouping related rows
    
    Args:
        rows_data: List of rows (each row is list of cell dicts)
        
    Returns:
        Fixed rows data
    """
    if not rows_data:
        return rows_data
    
    fixed_rows = []
    
    # Step 1: Remove empty rows
    non_empty_rows = []
    for row_data in rows_data:
        if any(cell.get('value', '').strip() for cell in row_data):
            non_empty_rows.append(row_data)
    
    if not non_empty_rows:
        return rows_data
    
    # Step 2: Process rows - create 2-column layout
    # Pattern: Label rows followed by value rows (or label:value in same row)
    i = 0
    while i < len(non_empty_rows):
        row_data = non_empty_rows[i]
        
        # Get first non-empty cell (label)
        first_cell = None
        for cell in row_data:
            if cell.get('value', '').strip():
                first_cell = cell
                break
        
        if not first_cell:
            i += 1
            continue
        
        label_text = first_cell['value'].strip()
        
        # Check if next row might be a value
        value_text = ''
        if i + 1 < len(non_empty_rows):
            next_row = non_empty_rows[i + 1]
            next_first_cell = None
            for cell in next_row:
                if cell.get('value', '').strip():
                    next_first_cell = cell
                    break
            
            if next_first_cell:
                next_text = next_first_cell['value'].strip()
                # If next row looks like a value (not a label pattern), use it
                # Labels typically end with ":", or are short, or contain keywords
                is_label = (':' in next_text or 
                           len(next_text.split()) <= 3 or
                           any(kw in next_text.lower() for kw in ['number', 'name', 'date', 'amount', 'id', 'url', 'status', 'channel']))
                
                if not is_label:
                    value_text = next_text
                    i += 1  # Skip next row as it's the value
        
        # Create new row with label in col 1, value in col 2
        new_row = copy.deepcopy(row_data)
        
        # Ensure we have at least 2 columns
        while len(new_row) < 2:
            new_cell = {
                'value': '',
                'row': first_cell['row'],
                'col': len(new_row) + 1,
                'cell': None
            }
            new_row.append(new_cell)
        
        # Set label in column 1
        new_row[0]['value'] = label_text
        new_row[0]['row'] = first_cell['row']
        new_row[0]['col'] = 1
        
        # Set value in column 2
        if value_text:
            new_row[1]['value'] = value_text
            new_row[1]['row'] = first_cell['row']
            new_row[1]['col'] = 2
        else:
            new_row[1]['value'] = ''
            new_row[1]['row'] = first_cell['row']
            new_row[1]['col'] = 2
        
        # Clear other columns
        for j in range(2, len(new_row)):
            new_row[j]['value'] = ''
        
        fixed_rows.append(new_row)
        i += 1
    
    return fixed_rows if fixed_rows else rows_data
