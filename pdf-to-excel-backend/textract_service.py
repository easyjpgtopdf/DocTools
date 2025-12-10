"""
AWS Textract service for extracting tables from PDF documents.
(Kept for fallback support - /api/pdf-to-excel endpoint)
"""

import os
import boto3
from botocore.exceptions import ClientError
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# Initialize Textract client
textract_client = None


def get_textract_client():
    """Initialize and return Textract client."""
    global textract_client
    if textract_client is None:
        textract_client = boto3.client(
            'textract',
            aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'),
            region_name=os.environ.get('AWS_REGION', 'us-east-1')
        )
    return textract_client


def analyze_document_with_tables(s3_bucket: str, s3_key: str) -> Dict[str, Any]:
    """
    Call AWS Textract AnalyzeDocument API with TABLES feature.
    
    Args:
        s3_bucket: S3 bucket name
        s3_key: S3 object key
    
    Returns:
        Textract response containing blocks with table data
    
    Raises:
        Exception: If Textract call fails
    """
    try:
        textract = get_textract_client()
        
        response = textract.analyze_document(
            Document={
                'S3Object': {
                    'Bucket': s3_bucket,
                    'Name': s3_key
                }
            },
            FeatureTypes=['TABLES']
        )
        
        return response
    except ClientError as e:
        error_code = e.response.get('Error', {}).get('Code', '')
        error_msg = e.response.get('Error', {}).get('Message', str(e))
        
        if error_code == 'InvalidParameterException':
            raise Exception(f"Invalid PDF file: {error_msg}")
        elif error_code == 'InvalidS3ObjectException':
            raise Exception(f"PDF file not found in S3: {error_msg}")
        elif error_code == 'UnsupportedDocumentException':
            raise Exception(f"Unsupported document format: {error_msg}")
        else:
            raise Exception(f"Textract analysis failed: {error_msg}")


def get_page_count(textract_response: Dict[str, Any]) -> int:
    """Extract page count from Textract response."""
    document_metadata = textract_response.get('DocumentMetadata', {})
    return document_metadata.get('Pages', 1)


def parse_textract_tables(textract_response: Dict[str, Any]) -> List[List[List[str]]]:
    """Parse Textract response to extract table data."""
    blocks = textract_response.get('Blocks', [])
    block_map = {block['Id']: block for block in blocks}
    
    tables = []
    for block in blocks:
        if block['BlockType'] == 'TABLE':
            table = parse_table_block(block, block_map)
            if table:
                tables.append(table)
    
    return tables


def parse_table_block(table_block: Dict, block_map: Dict) -> List[List[str]]:
    """Parse a single table block from Textract."""
    relationships = table_block.get('Relationships', [])
    cells = []
    
    for relationship in relationships:
        if relationship['Type'] == 'CHILD':
            for cell_id in relationship['Ids']:
                if cell_id in block_map:
                    cell_block = block_map[cell_id]
                    if cell_block['BlockType'] == 'CELL':
                        cells.append(cell_block)
    
    if not cells:
        return []
    
    max_row = max(cell.get('RowIndex', 1) for cell in cells)
    max_col = max(cell.get('ColumnIndex', 1) for cell in cells)
    grid = [['' for _ in range(max_col)] for _ in range(max_row)]
    
    for cell in cells:
        row_idx = cell.get('RowIndex', 1) - 1
        col_idx = cell.get('ColumnIndex', 1) - 1
        cell_text = extract_cell_text(cell, block_map)
        grid[row_idx][col_idx] = cell_text
    
    return grid


def extract_cell_text(cell_block: Dict, block_map: Dict) -> str:
    """Extract text content from a cell block."""
    relationships = cell_block.get('Relationships', [])
    text_parts = []
    
    for relationship in relationships:
        if relationship['Type'] == 'CHILD':
            for word_id in relationship['Ids']:
                if word_id in block_map:
                    word_block = block_map[word_id]
                    if word_block['BlockType'] == 'WORD':
                        text_parts.append(word_block.get('Text', ''))
    
    return ' '.join(text_parts).strip()


def create_excel_from_tables(tables: List[List[List[str]]], pdf_filename: str) -> bytes:
    """Create Excel workbook from parsed tables."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    base_name = pdf_filename.replace('.pdf', '').replace('.PDF', '')
    
    if not tables:
        sheet = workbook.create_sheet(title="No Tables")
        sheet['A1'] = "No tables detected in this PDF."
        sheet['A1'].font = Font(italic=True, color="666666")
    else:
        for idx, table in enumerate(tables):
            sheet_name = f"Table {idx + 1}" if len(tables) > 1 else base_name[:31]
            sheet = workbook.create_sheet(title=sheet_name)
            
            for row_idx, row in enumerate(table, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    
    # Auto-adjust column widths
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def convert_pdf_to_excel(s3_bucket: str, s3_key: str, pdf_filename: str):
    """Complete workflow: Analyze PDF with Textract and create Excel."""
    logger.info(f"Calling Textract for {s3_key}")
    textract_response = analyze_document_with_tables(s3_bucket, s3_key)
    page_count = get_page_count(textract_response)
    logger.info(f"PDF has {page_count} pages")
    
    logger.info("Parsing tables from Textract response")
    tables = parse_textract_tables(textract_response)
    logger.info(f"Found {len(tables)} tables")
    
    logger.info("Creating Excel workbook")
    excel_bytes = create_excel_from_tables(tables, pdf_filename)
    
    return excel_bytes, page_count, tables

